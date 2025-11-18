import re
import openpyxl
import pandas as pd

# Load workbook
wb = openpyxl.load_workbook("../FILES/PFE Participant List 2013 - 2024.xlsx", data_only=True)

# Storage for participants and events
participants = []
events = []

# Loop through all relevant sheets
for sheet_name in wb.sheetnames:

    ws = wb[sheet_name]
    current_event = None
    event_title = ""
    event_location_date = ""

    for row in ws.iter_rows(values_only=True):
        cell = row[0]

        if isinstance(cell, str) and "PFE" in cell:
            # Found new event
            match = re.match(r"(PFE\d{2}M\d)\s+(.*)", cell)
            if match:
                current_event = match.group(1)
                event_title = match.group(2)
                event_location_date = row[1] if row[1] else ""
                events.append({
                    "Event": current_event,
                    "Title": event_title,
                    "LocationDate": event_location_date
                })

        elif isinstance(cell, str) and current_event:
            match = re.match(r"\d+\.\s+(.*)", cell.strip())
            if match:
                full_entry = match.group(1).strip()

                # Try comma first
                if ',' in full_entry:
                    name, position = full_entry.split(',', 1)
                # Then en dash
                elif ' – ' in full_entry:
                    name, position = full_entry.split(' – ', 1)
                # Then regular dash
                elif ' - ' in full_entry:
                    name, position = full_entry.split(' - ', 1)
                else:
                    name = full_entry
                    position = ""

                participants.append({
                    "Event": current_event,
                    "Name": name.strip(),
                    "Position": position.strip(),
                    "ID": "",
                    "Grade": "",
                    "Country": sheet_name
                })

# Create DataFrames
df_participants = pd.DataFrame(participants)[
    ["Event", "Name", "Position", "ID", "Grade", "Country"]
]
df_events = pd.DataFrame(events)

# Save to Excel
with pd.ExcelWriter("FILES/final_result.xlsx", engine="openpyxl", mode='w') as writer:
    df_participants.to_excel(writer, sheet_name="Participants", index=False)
    df_events.to_excel(writer, sheet_name="Events", index=False)

print("File saved to FILES/final_result.xlsx")
