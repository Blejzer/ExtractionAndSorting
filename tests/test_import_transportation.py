from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import services.import_service as import_service


def _build_workbook_bytes(travel_value: str = "Bus") -> bytes:
    """Create a minimal workbook to exercise transportation parsing."""

    wb = Workbook()

    # Participants sheet with header cells used for metadata extraction
    ws = wb.active
    ws.title = "Participants"
    ws["A1"] = "E1 TITLE"
    ws["A2"] = "JUNE 1 - 3 - Zagreb"

    # ParticipantsLista table (positions lookup)
    ws_positions = wb.create_sheet("List")
    ws_positions.append(["Name (Latin)", "Position", "Phone", "Email"])
    ws_positions.append(["DOE, John", "Advisor", "123", "john@example.com"])
    tbl_positions = Table(displayName="ParticipantsLista", ref="A1:D2")
    tbl_positions.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_positions.add_table(tbl_positions)

    # MAIN ONLINE → ParticipantsList table with a conflicting transportation value
    ws_online = wb.create_sheet("MAIN ONLINE")
    online_columns = [
        "Name",
        "Middle name",
        "Last name",
        "Gender",
        "Date of Birth (DOB)",
        "Place Of Birth (POB)",
        "Country of Birth",
        "Citizenship(s)",
        "Email address",
        "Phone number",
        "Travelling document type",
        "Travelling document number",
        "Travelling document issuance date",
        "Travelling document expiry date",
        "Travelling document issued by",
        "Transportation",
        "Travelling from",
        "Returning to",
        "Diet restrictions",
        "Organization",
        "Unit",
        "Position",
        "Rank",
        "Authority",
        "Short professional biography",
        "Bank name",
        "IBAN",
        "IBAN Type",
        "SWIFT",
    ]
    ws_online.append(online_columns)
    ws_online.append([
        "John",
        "",
        "Doe",
        "male",
        datetime(1990, 5, 4),
        "City",
        "Country",
        "Country",
        "john@example.com",
        "123",
        "Passport",
        "X123",
        datetime(2018, 6, 1),
        datetime(2028, 6, 1),
        "Authority",
        "No",
        "Airplane",
        "Main City",
        "Home City",
        "None",
        "Org",
        "Unit",
        "Advisor",
        "Captain",
        "No",
        "Bio",
        "Bank",
        "IBAN123",
        "EURO",
        "SWIFT",
    ])
    last_col = get_column_letter(len(online_columns))
    tbl_online = Table(displayName="ParticipantsList", ref=f"A1:{last_col}2")
    tbl_online.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_online.add_table(tbl_online)

    # Country table where the attendee originates. Travel column contains the
    # canonical transportation value that must be preserved.
    ws_country = wb.create_sheet("Alb")
    ws_country.append(["Name and last name", "Travel", "Grade"])
    ws_country.append(["John Doe", travel_value, "10"])
    tbl_country = Table(displayName="tableAlb", ref="A1:C2")
    tbl_country.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_country.add_table(tbl_country)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_transportation_comes_from_country_table(tmp_path):
    workbook = _build_workbook_bytes("Bus")
    path = tmp_path / "transport.xlsx"
    path.write_bytes(workbook)

    result = import_service.parse_for_commit(str(path))
    attendees = result["attendees"]
    assert len(attendees) == 1

    attendee = attendees[0]
    assert attendee["transportation"] == "Bus"
    # Ensure the conflicting MAIN ONLINE value was ignored
    assert attendee["transportation"] != "Airplane"


def test_transportation_blanks_do_not_use_main_online(tmp_path):
    workbook = _build_workbook_bytes("")
    path = tmp_path / "transport_blank.xlsx"
    path.write_bytes(workbook)

    result = import_service.parse_for_commit(str(path))
    attendees = result["attendees"]
    assert len(attendees) == 1

    attendee = attendees[0]
    assert attendee["transportation"] == ""
    assert attendee["transportation"] != "Airplane"

