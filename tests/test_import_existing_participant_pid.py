from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import services.import_service_v2 as import_service


ONLINE_COLUMNS = [
    "Name",
    "Middle name",
    "Last name",
    "Gender",
    "Date of Birth (DOB)",
    "Place Of Birth (POB)",
    "Country of Birth",
    "Citizenship(s)",
    "Email address",
    "Phone number",
    "Traveling document type",
    "Traveling document number",
    "Traveling document issuance date",
    "Traveling document expiry date",
    "Traveling document issued by",
    "Returning to",
    "Diet restrictions",
    "Organization",
    "Unit",
    "Position",
    "Rank",
    "Authority",
    "Short professional biography",
    "Bank name",
    "IBAN",
    "IBAN Type",
    "SWIFT",
]


def _append_online_row(ws, data: dict) -> None:
    ws.append([data.get(col, "") for col in ONLINE_COLUMNS])


def _build_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws["A1"] = "E1 TITLE"
    ws["A2"] = "JUNE 1 - 3 - Zagreb"

    ws_cost = wb.create_sheet("COST Overview")
    ws_cost["B15"] = "1000"

    ws_list = wb.create_sheet("List")
    ws_list.append(["Name (Latin)", "Position", "Phone", "Email"])
    ws_list.append(["BAJIĆ BRALIĆ, Ana Marija", "Advisor", "123", "ana@example.com"])
    tbl_list = Table(displayName="ParticipantsLista", ref="A1:D2")
    tbl_list.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_list.add_table(tbl_list)

    ws_online = wb.create_sheet("MAIN ONLINE")
    ws_online.append(ONLINE_COLUMNS)
    _append_online_row(
        ws_online,
        {
            "Name": "Ana",
            "Middle name": "Marija",
            "Last name": "Bajic Bralic",
            "Gender": "female",
            "Date of Birth (DOB)": datetime(1973, 5, 25),
            "Place Of Birth (POB)": "Radac",
            "Country of Birth": "Kosovo, Europe & Eurasia, World",
            "Citizenship(s)": "Kosovo, Europe & Eurasia",
            "Email address": "ana@example.com",
            "Phone number": "123",
            "Traveling document type": "Passport",
            "Traveling document number": "P01415451",
            "Traveling document issuance date": datetime(2019, 3, 27),
            "Traveling document expiry date": datetime(2029, 3, 26),
            "Traveling document issued by": "Republic of Kosovo",
            "Returning to": "Pristina",
            "Diet restrictions": "No pork, no chilli",
            "Organization": "Prosecution System",
            "Unit": "Peja Basic Prosecutor's Office",
            "Position": "Advisor",
            "Rank": "Chief prosecutor",
            "Authority": "Yes",
            "Short professional biography": "bio",
            "Bank name": "BANKA KOMBETARE TREGTARE KOSOVE SHA",
            "IBAN": "XK051920315886321195",
            "IBAN Type": "EURO",
            "SWIFT": "NCBA XK PR",
        },
    )
    last_col = get_column_letter(len(ONLINE_COLUMNS))
    tbl_online = Table(displayName="ParticipantsList", ref=f"A1:{last_col}2")
    tbl_online.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_online.add_table(tbl_online)

    ws_country = wb.create_sheet("Cro")
    ws_country.append([
        "Name and Last Name",
        "Grade (0 - BL, 1 - Pass, 2 - Excel)",
    ])
    ws_country.append(["BAJIĆ BRALIĆ, Ana Marija", ""])
    ws_country.append(["TOTAL", ""])
    tbl_country = Table(displayName="tableCro", ref="A1:B3")
    tbl_country.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_country.add_table(tbl_country)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

class DummyParticipant:
    def __init__(self, pid: str) -> None:
        self.pid = pid


class DummyRepo:
    def __init__(self) -> None:
        self.calls: list[tuple[str, datetime | None, str]] = []

    def find_by_name_dob_and_representing_country_cid(
        self, *, name: str, dob: datetime | None, representing_country: str
    ):
        self.calls.append((name, dob, representing_country))
        return DummyParticipant("P9999")


def test_existing_participant_pid_is_attached(monkeypatch, tmp_path):
    repo = DummyRepo()
    monkeypatch.setattr(import_service, "_participant_repo", repo)
    monkeypatch.setattr(
        import_service,
        "resolve_country_flexible",
        lambda raw: {"cid": "C117", "country": raw} if raw else None,
    )
    monkeypatch.setattr(
        import_service,
        "get_country_cid_by_name",
        lambda raw: "C200" if raw else None,
    )

    content = _build_workbook_bytes()
    path = tmp_path / "pid-lookup.xlsx"
    path.write_bytes(content)

    result = import_service.parse_for_commit(str(path))

    attendee = result["attendees"][0]
    assert attendee["pid"] == "P9999"
    assert repo.calls
    name, dob, country = repo.calls[0]
    assert name == "Ana Marija BAJIĆ BRALIĆ"
    assert country == "C200"
    assert isinstance(dob, datetime)

