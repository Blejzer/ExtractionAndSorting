from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import pytest

from domain.models.participant import Gender
import services.import_service_v2 as import_service


ONLINE_COLUMNS = [
    "Name",
    "Middle name",
    "Last name",
    "Gender",
    "Date of Birth (DOB)",
    "Place Of Birth (POB)",
    "Country of Birth",
    "Citizenship(s)",
    "Email address",
    "Phone number",
    "Transportation",
    "Transportation (Other)",
    "Traveling from",
    "Returning to",
    "Diet restrictions",
    "Organization",
    "Unit",
    "Position",
    "Rank",
    "Authority",
    "Short professional biography",
    "Bank name",
    "IBAN",
    "IBAN Type",
    "SWIFT",
]


def _participant_record(**overrides):
    record = {
        "pid": "P001",
        "representing_country": "HR",
        "gender": "Male",
        "grade": "1",
        "name": "John Doe",
        "dob": datetime(1990, 1, 1),
        "pob": "Zagreb",
        "birth_country": "HR",
        "citizenships": ["HR"],
    }
    record.update(overrides)
    return record


def _append_online_row(ws, data: dict) -> None:
    ws.append([data.get(col, "") for col in ONLINE_COLUMNS])


def _workbook_bytes_with_gender(gender: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws["A1"] = "E1 TITLE"
    ws["A2"] = "JUNE 1 - 3 - Zagreb"

    ws_list = wb.create_sheet("List")
    ws_list.append(["Name (Latin)", "Position", "Phone", "Email"])
    ws_list.append(["DOE, John", "Analyst", "+385123", "john@example.com"])
    tbl_list = Table(displayName="ParticipantsLista", ref="A1:D2")
    tbl_list.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_list.add_table(tbl_list)

    ws_online = wb.create_sheet("MAIN ONLINE")
    ws_online.append(ONLINE_COLUMNS)
    _append_online_row(
        ws_online,
        {
            "Name": "John",
            "Middle name": "",
            "Last name": "Doe",
            "Gender": gender,
            "Date of Birth (DOB)": datetime(1990, 1, 1),
            "Place Of Birth (POB)": "Zagreb",
            "Country of Birth": "Croatia, Europe & Eurasia, World",
            "Citizenship(s)": "Croatia, Europe & Eurasia",
            "Email address": "john@example.com",
            "Phone number": "+385123",
            "Transportation": "Flight",
            "Transportation (Other)": "",
            "Traveling from": "Zagreb",
            "Returning to": "Zagreb",
            "Diet restrictions": "None",
            "Organization": "Ministry",
            "Unit": "Unit",
            "Position": "Analyst",
            "Rank": "Officer",
            "Authority": "Yes",
            "Short professional biography": "Bio",
            "Bank name": "Bank",
            "IBAN": "HR123",
            "IBAN Type": "EURO",
            "SWIFT": "SWIFT",
        },
    )
    last_col = get_column_letter(len(ONLINE_COLUMNS))
    tbl_online = Table(displayName="ParticipantsList", ref=f"A1:{last_col}2")
    tbl_online.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_online.add_table(tbl_online)

    ws_country = wb.create_sheet("Cro")
    ws_country.append(["Name and Last Name", "Grade"])
    ws_country.append(["DOE, John", "1"])
    ws_country.append(["TOTAL", ""])
    tbl_country = Table(displayName="tableCro", ref="A1:B3")
    tbl_country.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_country.add_table(tbl_country)

    ws_cost = wb.create_sheet("COST Overview")
    ws_cost["B15"] = ""

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@pytest.mark.parametrize(
    "raw_gender, expected",
    [
        ("Mr", Gender.male),
        ("Mrs", Gender.female),
        ("Ms", Gender.female),
    ],
)
def test_build_participant_normalizes_gender_titles(raw_gender, expected):
    participant = import_service._build_participant_from_record(
        _participant_record(gender=raw_gender)
    )

    assert participant is not None
    assert participant.gender == expected


@pytest.mark.parametrize(
    "raw_gender, expected",
    [
        ("Mr", "Male"),
        ("Ms", "Female"),
    ],
)
def test_parse_for_commit_normalizes_gender_in_attendees(tmp_path, raw_gender, expected):
    workbook_path = tmp_path / "import.xlsx"
    workbook_path.write_bytes(_workbook_bytes_with_gender(raw_gender))

    result = import_service.parse_for_commit(str(workbook_path))

    assert result["attendees"], "Expected attendees to be parsed"
    assert result["attendees"][0]["gender"] == expected


def test_parse_for_commit_attaches_existing_pid(tmp_path, monkeypatch):
    workbook_path = tmp_path / "import_existing.xlsx"
    workbook_path.write_bytes(_workbook_bytes_with_gender("Male"))

    captured: dict = {}

    def _fake_exists(name_display, country_name, dob_iso=None):
        captured["call"] = (name_display, country_name, dob_iso)
        return True, {"pid": "P7777"}

    monkeypatch.setattr(import_service, "_participant_exists", _fake_exists)
    monkeypatch.setattr(import_service, "resolve_country_flexible", lambda value: {"cid": "HR", "country": "Croatia"})
    monkeypatch.setattr(import_service, "get_country_cid_by_name", lambda value: "HR")

    result = import_service.parse_for_commit(str(workbook_path))

    assert result["attendees"], "Expected attendees to be parsed"

    attendee = result["attendees"][0]
    assert attendee["pid"] == "P7777"

    preview_participant = result["preview"]["participants"][0]
    assert preview_participant["pid"] == "P7777"

    assert captured["call"][2] == attendee.get("dob")
