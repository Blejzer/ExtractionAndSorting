from io import BytesIO

from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import services.import_service as import_service


ONLINE_COLUMNS = [
    "Name",
    "Middle name",
    "Last name",
    "Gender",
    "Date of Birth (DOB)",
    "Place Of Birth (POB)",
    "Country of Birth",
    "Citizenship(s)",
    "Email address",
    "Phone number",
    "Travelling document type",
    "Travelling document number",
    "Travelling document issuance date",
    "Travelling document expiry date",
    "Travelling document issued by",
    "Do you require Visa to travel to Croatia",
    "Returning to",
    "Diet restrictions",
    "Organization",
    "Unit",
    "Position",
    "Rank",
    "Authority",
    "Short professional biography",
    "Bank name",
    "IBAN",
    "IBAN Type",
    "SWIFT",
]


def _append_online_row(ws, data: dict) -> None:
    ws.append([data.get(col, "") for col in ONLINE_COLUMNS])


def _build_workbook_bytes() -> bytes:
    """Construct a minimal workbook containing the required tables."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws["A1"] = "E1 TITLE"
    ws["A2"] = "JUNE 1 - 3 - Zagreb"

    # ParticipantsLista with position/phone/email
    ws_list = wb.create_sheet("List")
    ws_list.append(["Name (Latin)", "Position", "Phone", "Email"])
    ws_list.append(["BAJIĆ BRALIĆ, Ana Marija", "Advisor", "123", "ana@example.com"])
    tbl_list = Table(displayName="ParticipantsLista", ref="A1:D2")
    tbl_list.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_list.add_table(tbl_list)

    # MAIN ONLINE → ParticipantsList (minimal)
    ws_online = wb.create_sheet("MAIN ONLINE")
    ws_online.append(ONLINE_COLUMNS)
    _append_online_row(
        ws_online,
        {
            "Name": "Ana",
            "Middle name": "Marija",
            "Last name": "Bajic Bralic",
            "Gender": "female",
            "Date of Birth (DOB)": datetime(1973, 5, 25),
            "Place Of Birth (POB)": "Radac",
            "Country of Birth": "Kosovo, Europe & Eurasia, World",
            "Citizenship(s)": "Kosovo, Europe & Eurasia",
            "Email address": "ana@example.com",
            "Phone number": "123",
            "Travelling document type": "Passport",
            "Travelling document number": "P01415451",
            "Travelling document issuance date": datetime(2019, 3, 27),
            "Travelling document expiry date": datetime(2029, 3, 26),
            "Travelling document issued by": "Republic of Kosovo",
            "Do you require Visa to travel to Croatia": "No",
            "Returning to": "Pristina",
            "Diet restrictions": "No pork, no chilli",
            "Organization": "Prosecution System",
            "Unit": "Peja Basic Prosecutor's Office",
            "Position": "Advisor",
            "Rank": "Chief prosecutor",
            "Authority": "Yes",
            "Short professional biography": "bio",
            "Bank name": "BANKA KOMBETARE TREGTARE KOSOVE SHA",
            "IBAN": "XK051920315886321195",
            "IBAN Type": "EURO",
            "SWIFT": "NCBA XK PR",
        },
    )
    last_col = get_column_letter(len(ONLINE_COLUMNS))
    tbl_online = Table(displayName="ParticipantsList", ref=f"A1:{last_col}2")
    tbl_online.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_online.add_table(tbl_online)

    # Country table with the attendee
    ws_country = wb.create_sheet("Cro")
    ws_country.append(["Name and last name", "Grade"])
    ws_country.append(["BAJIĆ BRALIĆ, Ana Marija", ""])
    ws_country.append(["TOTAL", ""])
    tbl_country = Table(displayName="tableCro", ref="A1:B3")
    tbl_country.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_country.add_table(tbl_country)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _build_workbook_bytes_middle_name_variant() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws["A1"] = "E1 TITLE"
    ws["A2"] = "JUNE 1 - 3 - Zagreb"

    ws_list = wb.create_sheet("List")
    ws_list.append(["Name (Latin)", "Position", "Phone", "Email"])
    ws_list.append([
        "STEPANOVIĆ, Aleksandar",
        "Inspector",
        "+381648923499",
        "alekstepanovic@hotmail.com",
    ])
    tbl_list = Table(displayName="ParticipantsLista", ref="A1:D2")
    tbl_list.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_list.add_table(tbl_list)

    ws_online = wb.create_sheet("MAIN ONLINE")
    ws_online.append(ONLINE_COLUMNS)
    _append_online_row(
        ws_online,
        {
            "Name": "Aleksandar",
            "Middle name": "Nikola",
            "Last name": "Stepanovic",
            "Gender": "male",
            "Date of Birth (DOB)": datetime(1990, 1, 2),
            "Place Of Birth (POB)": "Belgrade",
            "Country of Birth": "Serbia",
            "Citizenship(s)": "Serbia",
            "Email address": "alekstepanovic@hotmail.com",
            "Phone number": "+381648923499",
            "Travelling document type": "Passport",
            "Travelling document number": "S1234567",
            "Travelling document issuance date": datetime(2020, 5, 14),
            "Travelling document expiry date": datetime(2030, 5, 14),
            "Travelling document issued by": "MUP R SERBIA, PU IN VRANJE",
            "Do you require Visa to travel to Croatia": "No",
            "Returning to": "Serbia",
            "Diet restrictions": "NO RESTRICTIONS",
            "Organization": "MINISTRY OF INTERIOR AFFAIRS",
            "Unit": "Combating Environmental Crime Department",
            "Position": "CIU Inspector",
            "Rank": "Manager",
            "Authority": "No",
            "Short professional biography": "bio",
            "Bank name": "Banka poštanska štedionica",
            "IBAN": "RS35200075000001314017",
            "IBAN Type": "EURO",
            "SWIFT": "SBPORSBG",
        },
    )
    last_col = get_column_letter(len(ONLINE_COLUMNS))
    tbl_online = Table(displayName="ParticipantsList", ref=f"A1:{last_col}2")
    tbl_online.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_online.add_table(tbl_online)

    ws_country = wb.create_sheet("Ser")
    ws_country.append(["Name and last name", "Grade"])
    ws_country.append(["STEPANOVIĆ, Aleksandar", 1])
    ws_country.append(["TOTAL", ""])
    tbl_country = Table(displayName="tableSer", ref="A1:B3")
    tbl_country.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws_country.add_table(tbl_country)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_bajic_bralic_lookup(tmp_path):
    content = _build_workbook_bytes()
    path = tmp_path / "sample.xlsx"
    with open(path, "wb") as fh:
        fh.write(content)

    result = import_service.parse_for_commit(str(path))
    attendees = result["attendees"]
    assert len(attendees) == 1
    attendee = attendees[0]
    assert attendee["name"] == "Ana Marija BAJIĆ BRALIĆ"
    assert attendee["position"] == "Advisor"
    assert attendee["phone"] == "123"
    assert attendee["email"] == "ana@example.com"
    assert attendee["gender"] == "Female"
    assert attendee["dob"] == "1973-05-25"
    assert attendee["pob"] == "Radac"
    assert attendee["birth_country"] == "Kosovo, Europe & Eurasia"
    assert attendee["citizenships"] == ["C117"]
    assert attendee["travel_doc_type"] == "Passport"
    assert attendee["travel_doc_number"] == "P01415451"
    assert attendee["travel_doc_issue_date"] == "2019-03-27"
    assert attendee["travel_doc_expiry_date"] == "2029-03-26"
    assert attendee["travel_doc_issued_by"] == "Republic of Kosovo"
    assert attendee["returning_to"] == "Pristina"
    assert attendee["diet_restrictions"] == "No pork, no chilli"
    assert attendee["organization"] == "Prosecution System"
    assert attendee["unit"] == "Peja Basic Prosecutor's Office"
    assert attendee["rank"] == "Chief prosecutor"
    assert attendee["intl_authority"] is True
    assert attendee["bio_short"] == "bio"
    assert attendee["bank_name"] == "BANKA KOMBETARE TREGTARE KOSOVE SHA"
    assert attendee["iban"] == "XK051920315886321195"
    assert attendee["iban_type"] == "EURO"
    assert attendee["swift"] == "NCBA XK PR"
    assert attendee["grade"] == 1

    # Ensure debug-only data is not present by default
    assert "initial_attendees" not in result


def test_main_online_middle_name_is_optional(tmp_path):
    content = _build_workbook_bytes_middle_name_variant()
    path = tmp_path / "middle-name.xlsx"
    with open(path, "wb") as fh:
        fh.write(content)

    result = import_service.parse_for_commit(str(path))
    attendees = result["attendees"]
    assert len(attendees) == 1

    attendee = attendees[0]
    assert attendee["name"] == "Aleksandar STEPANOVIĆ"
    assert attendee["position"] == "Inspector"
    assert attendee["phone"] == "+381648923499"
    assert attendee["email"] == "alekstepanovic@hotmail.com"
    assert attendee["gender"] == "Male"
    assert attendee["dob"] == "1990-01-02"
    assert attendee["pob"] == "Belgrade"
    assert attendee["birth_country"] == "Serbia"
    assert attendee["citizenships"] == ["C194"]
    assert attendee["travel_doc_number"] == "S1234567"
    assert attendee["iban"] == "RS35200075000001314017"

