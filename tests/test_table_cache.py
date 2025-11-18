import pandas as pd
import openpyxl

from services import import_service_v2
from services.xlsx_tables_inspector import TableRef


def _make_workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws["A1"].value = "Name"
    ws["B1"].value = "Grade"
    ws["A2"].value = "Alice"
    ws["B2"].value = 1
    path = tmp_path / "cache.xlsx"
    wb.save(path)
    return path


def test_extract_all_tables_caches_duplicate_tables(monkeypatch, tmp_path):
    path = _make_workbook(tmp_path)

    table = TableRef(
        name="tableAlb",
        name_norm="tablealb",
        sheet_title="Participants",
        ref="A1:B2",
        table_xml_path="xl/tables/table1.xml",
    )

    tables = [table, table]
    calls = {"count": 0}

    def fake_builder(ws, table_ref):
        calls["count"] += 1
        return pd.DataFrame({"Name": ["Alice"], "Grade": [1]})

    monkeypatch.setattr(import_service_v2, "_build_table_dataframe", fake_builder)

    cache = import_service_v2._extract_all_tables(str(path), tables)

    assert calls["count"] == 1
    df = cache.get_df(table)
    assert list(df.columns) == ["Name", "Grade"]
    assert df.iloc[0]["Name"] == "Alice"
