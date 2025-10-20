import os
from io import BytesIO

import json
import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


from app import create_app


def _build_workbook_bytes(valid: bool) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws["A1"] = "E1 Title"
    ws["A2"] = "2024"

    if valid:
        ws_list = wb.create_sheet("List")
        ws_list.append(["Name", "Position"])
        ws_list.append(["Doe John", "Leader"])
        tbl = Table(displayName="ParticipantsLista", ref="A1:B2")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws_list.add_table(tbl)

        ws_country = wb.create_sheet("Alb")
        ws_country.append(["Name and last name", "Grade"])
        ws_country.append(["John Doe", "10"])
        tbl2 = Table(displayName="tableAlb", ref="A1:B2")
        tbl2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws_country.add_table(tbl2)

        # MAIN ONLINE sheet with ParticipantsList table for enrichment
        ws_online = wb.create_sheet("MAIN ONLINE")
        online_cols = [
            "Name", "Middle name", "Last name", "Gender",
            "Travelling document type", "Travelling document number",
        ]
        ws_online.append(online_cols)
        ws_online.append([
            "John", "", "Doe", "male",
            "Passport", "X123456",
        ])
        last_col = get_column_letter(len(online_cols))
        tbl_online = Table(displayName="ParticipantsList", ref=f"A1:{last_col}2")
        tbl_online.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws_online.add_table(tbl_online)


    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@pytest.fixture
def client(tmp_path):
    app = create_app()
    app.config["TESTING"] = True
    app.config["UPLOADS_DIR"] = tmp_path
    os.makedirs(tmp_path, exist_ok=True)
    with app.test_client() as client:
        with client.session_transaction() as sess:
            sess["username"] = "tester"
        yield client


def test_upload_valid_file(client):
    data = {"file": (BytesIO(_build_workbook_bytes(True)), "sample.xlsx")}
    resp = client.post("/import", data=data, content_type="multipart/form-data")
    assert resp.status_code == 200
    assert b"File is OK" in resp.data


def test_upload_invalid_file(client):
    data = {"file": (BytesIO(_build_workbook_bytes(False)), "bad.xlsx")}
    resp = client.post("/import", data=data, content_type="multipart/form-data")
    assert resp.status_code == 200
    assert b"File is not formatted correctly" in resp.data


def test_proceed_and_discard(client, tmp_path):
    content = _build_workbook_bytes(True)
    path = tmp_path / "sample.xlsx"
    with open(path, "wb") as fh:
        fh.write(content)

    resp = client.post("/import/proceed", data={"filename": "sample.xlsx"})
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/import/preview/sample.preview.json")

    preview_resp = client.get("/import/preview/sample.preview.json")
    assert preview_resp.status_code == 200
    assert b"Participants" in preview_resp.data


    # preview JSON should contain enriched fields from ParticipantsList
    preview_path = tmp_path / "sample.preview.json"
    assert preview_path.exists()
    with open(preview_path, "r", encoding="utf-8") as fh:
        preview = json.load(fh)
    participant = preview["participants"][0]
    assert participant["gender"] == "Male"
    assert participant["travel_doc_number"] == "X123456"
    assert preview["participant_events"] == []

    # recreate file for discard test
    with open(path, "wb") as fh:
        fh.write(content)
    assert path.exists()
    resp = client.post("/import/discard", data={"filename": "sample.xlsx"})
    assert resp.status_code == 302
    assert not path.exists()
