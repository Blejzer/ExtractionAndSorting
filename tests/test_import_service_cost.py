import pytest
from openpyxl import Workbook

from services.import_service import _parse_cost_value, _read_event_header_block


def test_parse_cost_value_handles_various_formats():
    assert _parse_cost_value(12500) == pytest.approx(12500.0)
    assert _parse_cost_value("12,345.67") == pytest.approx(12345.67)
    assert _parse_cost_value("12.345,67") == pytest.approx(12345.67)
    assert _parse_cost_value("EUR 9.876,54") == pytest.approx(9876.54)
    assert _parse_cost_value(" ") is None
    assert _parse_cost_value(None) is None
    assert _parse_cost_value(True) is None


def test_read_event_header_block_reads_cost_case_insensitive(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "participants"
    ws["A1"] = "PFE25M2 Example Event"
    ws["A2"] = "JUNE 1 - 5 - City, COUNTRY"

    cost_ws = wb.create_sheet("Cost overview")
    cost_ws["B15"] = "123,456.78 EUR"

    path = tmp_path / "preview.xlsx"
    wb.save(path)

    (
        eid,
        title,
        start_date,
        end_date,
        place,
        country,
        cost,
    ) = _read_event_header_block(str(path))

    assert eid == "PFE25M2"
    assert title == "Example Event"
    assert start_date is not None
    assert end_date is not None
    assert place == "City"
    assert cost == pytest.approx(123456.78)
