# services/import_service_v2.py
"""
================================================================================
Excel Import and Validation Service (Refactored for Readability)
================================================================================

Purpose:
--------
This module parses complex Excel workbooks used for event and participant data.
It validates required sheets and tables, extracts structured data, and builds
preview payloads for database import—without performing any writes.

Functional Sections:
--------------------
1. Configuration & Constants
2. Name / String Normalization Helpers
3. XML Extraction and Parsing Utilities
4. Data Coercion Helpers (dates, enums, booleans)
5. Object Builders (Event, Participant, EventParticipant)
6. Serialization for Preview
7. Lookups (ParticipantsLista, MAIN ONLINE tables)
8. Parsing Logic (country tables, enrichment, payload creation)
9. Validation & Inspection (Excel structure, preview)
10. Workbook/Table Utilities & DB Lookups

Notes:
------
• All logic and behavior are identical to the original `import_service.py`.
• Debug prints remain intact (toggle with `DEBUG_PRINT`).
• No DB writes occur here; only structure, readability, and order improved.
• Type hints will be added in the optimization phase.

================================================================================
"""

# === Standard Library Imports ===
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime, UTC
from typing import Dict, List, Optional, Any

# === Third-Party Imports ===
import openpyxl
import pandas as pd
from openpyxl.utils import range_boundaries

from config.database import mongodb
from config.settings import DEBUG_PRINT, REQUIRE_PARTICIPANTS_LIST
from domain.models.event import Event, EventType
from domain.models.event_participant import DocType, EventParticipant
from domain.models.participant import Grade, Participant
from repositories.participant_repository import ParticipantRepository
from services.xlsx_tables_inspector import list_tables, TableRef
from utils.country_resolver import COUNTRY_TABLE_MAP, resolve_country_flexible, get_country_cid_by_name, \
    _split_multi_country
from utils.dates import MONTHS, normalize_dob, coerce_datetime, date_to_iso
# === Internal Imports ===
from utils.excel import WorkbookCache
from utils.excel import _norm_tablename, get_mapping
from utils.names import (
    _name_key,
    _name_key_from_raw,
    _split_name_variants,
    _to_app_display_name,
)
from utils.normalize_phones import normalize_phone
from utils.participants import _normalize_gender, lookup
from utils.translation import translate

# ==============================================================================
# 1. Configuration & Constants
# ==============================================================================

try:  # pragma: no cover - exercised indirectly via repo lookups
    _participant_repo: Optional[ParticipantRepository] = ParticipantRepository()
except Exception:  # pragma: no cover - allow parsing when DB is unavailable
    _participant_repo = None  # type: ignore

# ==============================================================================
# 2. Custom XML Extraction and Parsing Utilities
# ==============================================================================

def _strip_xml_tag(tag: str) -> str:
    """Remove namespace from an XML tag."""
    return tag.split("}", 1)[1] if "}" in tag else tag


def _element_to_flat_dict(elem: ET.Element, prefix: str = "") -> Dict[str, str]:
    """
    Recursively flatten an XML element into a key-value dict.
    Nested elements become keys joined with underscores.

    Example:
        <participant>
            <name>John</name>
            <organization>MOI</organization>
        </participant>
        → {"participant_name": "John", "participant_organization": "MOI"}
    """
    children = list(elem)
    if not children:
        key = prefix or _strip_xml_tag(elem.tag)
        return {key: (elem.text or "").strip()}

    data: Dict[str, str] = {}
    for child in children:
        child_tag = _strip_xml_tag(child.tag)
        child_prefix = f"{prefix}_{child_tag}" if prefix else child_tag
        child_data = _element_to_flat_dict(child, child_prefix)
        for key, value in child_data.items():
            if not value:
                continue
            if key in data and data[key]:
                data[key] = f"{data[key]}; {value}"
            else:
                data[key] = value
    return data


def _collect_custom_xml_records(path: str) -> Optional[Dict[str, List[Dict[str, str]]]]:
    """
    Collect embedded CustomXML parts from an Excel .xlsx file.

    Returns:
        dict mapping category → list[record_dict] or None if no records.
        Example keys: 'participant', 'event', 'participant_event'.
    """
    try:
        with zipfile.ZipFile(path) as zf:
            names = [
                n for n in zf.namelist()
                if n.startswith("customXml/") and n.endswith(".xml")
            ]
            if not names:
                return None

            collected = {
                "participant": [],
                "event": [],
                "participant_event": [],
            }

            for name in names:
                try:
                    root = ET.fromstring(zf.read(name))
                except ET.ParseError:
                    if DEBUG_PRINT:
                        print(f"[CUSTOM-XML] Failed to parse {name}")
                    continue

                # Depth-first traversal
                stack = [root]
                while stack:
                    node = stack.pop()
                    tag = _strip_xml_tag(node.tag)
                    if tag in collected:
                        collected[tag].append(_element_to_flat_dict(node))
                    stack.extend(list(node))

            if not any(collected.values()):
                return None

            return {
                "participants": collected["participant"],
                "events": collected["event"],
                "participant_events": collected["participant_event"],
            }
    except (zipfile.BadZipFile, FileNotFoundError):
        return None
# ==============================================================================
# 4. Data Coercion and Normalization Helpers  (REPLACED / OPTIMIZED)
# ==============================================================================

from zoneinfo import ZoneInfo

EU_TZ = ZoneInfo("Europe/Zagreb")  # Used for all datetime coercion

# --- Internal micro-helpers ----------------------------------------------------

_BOOL_MAP = {
    "yes": True,
    "no": False,
    "true": True,
    "false": False,
}

def _as_str_or_empty(obj: object) -> str:
    """Fast, null-safe conversion to stripped string."""
    return str(obj).strip() if obj is not None else ""


# --- Boolean ------------------------------------------------------------------

def _parse_bool_value(value: object) -> Optional[bool]:
    """
    Accept only True/False or case-insensitive Yes/No.
    Everything else returns None.
    """
    if isinstance(value, bool):
        return value
    s = _as_str_or_empty(value).lower()
    return _BOOL_MAP.get(s)


# --- Grade --------------------------------------------------------------------

def _coerce_grade_value(value: object) -> int:
    """
    Accept only integers 0, 1, 2 (Normal=1 default).
    Any invalid or out-of-range value → 1.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 1
    try:
        iv = int(float(value))
        return iv if iv in (0, 1, 2) else 1
    except Exception:
        s = _as_str_or_empty(value)
        if s.lower() == "normal":
            return 1
        return 1


# --- EventType coercion -------------------------------------------------------

def _coerce_event_type(value: object) -> Optional[EventType]:
    """
    Coerce a raw value to EventType if possible.
    Accepts EventType or case-insensitive string.
    """
    if value is None:
        return None
    if isinstance(value, EventType):
        return value
    s = _as_str_or_empty(value)
    if not s:
        return None
    for variant in (s, s.title(), s.upper()):
        try:
            return EventType(variant)
        except ValueError:
            continue
    return None

# ==============================================================================
# 5. Object Builders (Event, Participant, EventParticipant)
# ==============================================================================

def _build_event_from_record(record: Dict[str, str]) -> Event:
    """
    Build an Event model instance from a raw record dictionary.
    Handles basic coercions for dates, cost, and event type.
    """
    start_date = coerce_datetime(record.get("start_date"), tzinfo=EU_TZ)
    end_date = coerce_datetime(record.get("end_date"), tzinfo=EU_TZ)

    # Cost coercion
    cost_val: Optional[float] = None
    cost_raw = record.get("cost")
    if cost_raw not in (None, ""):
        try:
            cost_val = float(str(cost_raw).strip())
        except ValueError:
            cost_val = None

    event_type = _coerce_event_type(record.get("type"))

    return Event(
        eid=str(record.get("eid", "")),
        title=str(record.get("title", "")),
        start_date=start_date,
        end_date=end_date,
        place=str(record.get("place", "")),
        country=record.get("country") or None,
        type=event_type,
        cost=cost_val,
    )


def _build_participant_from_record(record: Dict[str, str]) -> Optional[Participant]:
    """
    Build a Participant model instance from a raw record dictionary.
    Handles normalization of gender, grade, DOB, and boolean fields.
    """
    data: Dict[str, Any] = dict(record)

    normalized_gender = _normalize_gender(data.get("gender"))
    if normalized_gender is not None:
        data["gender"] = normalized_gender

    data["grade"] = _coerce_grade_value(data.get("grade"))
    dob_dt = normalize_dob(data.get("dob"))
    if dob_dt:
        data["dob"] = dob_dt

    for field in ["intl_authority"]:
        if field in data:
            val = _parse_bool_value(data[field])
            if val is not None:
                data[field] = val

    try:
        return Participant.model_validate(data)
    except Exception as exc:
        if DEBUG_PRINT:
            print(f"[CUSTOM-XML] Failed to build Participant: {exc}")
        return None


def _build_participant_event_from_record(record: Dict[str, str]) -> Optional[EventParticipant]:
    """
    Build an EventParticipant model instance from a raw record dictionary.
    Ensures date coercion for travel documents.
    """
    data: Dict[str, Any] = dict(record)
    for key in ("travel_doc_issue_date", "travel_doc_expiry_date"):
        if key in data:
            data[key] = coerce_datetime(data.get(key), tzinfo=EU_TZ)

    try:
        return EventParticipant.model_validate(data)
    except Exception as exc:
        if DEBUG_PRINT:
            print(f"[CUSTOM-XML] Failed to build EventParticipant: {exc}")
        return None


# ==============================================================================
# 6. Serialization Helpers (Preview / Merging)  (REPLACED / OPTIMIZED)
# ==============================================================================

def _serialize_model_for_preview(obj, enum_fields: tuple = (), ensure_int_fields: tuple = ()):
    """
    Generic serializer for Pydantic models:
      • Converts Enums → .value
      • Keeps datetime fields as datetime (Mongo-compatible)
      • Casts selected fields to int (e.g., grade)
    """
    if obj is None:
        return {}

    data = obj.model_dump(exclude_none=True)
    out: Dict[str, Any] = {}

    for key, val in data.items():
        if key in enum_fields and hasattr(val, "value"):
            out[key] = val.value
        elif isinstance(val, datetime):
            # keep native datetime (already EU tz)
            out[key] = val
        elif key in ensure_int_fields:
            try:
                out[key] = int(val)
            except Exception:
                out[key] = 1
        else:
            out[key] = val
    return out


def _serialize_event_for_preview(event: Optional[Event]) -> Dict[str, Any]:
    """Serialize Event → dict with datetime kept for Mongo."""
    return _serialize_model_for_preview(event, enum_fields=("type",))


def _serialize_participant_for_preview(participant: Participant) -> Dict[str, Any]:
    """Serialize Participant → dict with enums and datetimes preserved."""
    return _serialize_model_for_preview(
        participant,
        enum_fields=("gender",),
        ensure_int_fields=("grade",),
    )


def _serialize_participant_event_for_preview(ep: EventParticipant) -> Dict[str, Any]:
    """Serialize EventParticipant → dict with enums and datetimes preserved."""
    return _serialize_model_for_preview(
        ep,
        enum_fields=("transportation", "travel_doc_type", "iban_type"),
    )


def _merge_attendee_preview(participant: Participant, ep: EventParticipant) -> Dict[str, Any]:
    """
    Combine Participant and EventParticipant data into a single attendee preview record.
    Keeps all datetime objects (Mongo-ready) and converts Enums to .value strings.
    """
    transportation = (
        ep.transportation.value if hasattr(ep.transportation, "value") else ep.transportation
    )
    travel_doc_type = (
        ep.travel_doc_type.value if hasattr(ep.travel_doc_type, "value") else ep.travel_doc_type
    )
    iban_type = (
        ep.iban_type.value if hasattr(ep.iban_type, "value") else ep.iban_type
    )

    attendee: Dict[str, Any] = {
        "pid": participant.pid,
        "name": participant.name,
        "representing_country": participant.representing_country,
        "gender": participant.gender.value if hasattr(participant.gender, "value") else participant.gender,
        "grade": int(participant.grade),
        "dob": participant.dob,  # keep datetime
        "pob": participant.pob,
        "birth_country": participant.birth_country,
        "citizenships": participant.citizenships or [],
        "email": participant.email,
        "phone": participant.phone,
        "diet_restrictions": participant.diet_restrictions,
        "organization": participant.organization,
        "unit": participant.unit,
        "position": participant.position,
        "rank": participant.rank,
        "intl_authority": participant.intl_authority,
        "bio_short": participant.bio_short,
        "event_id": ep.event_id,
        "participant_id": ep.participant_id,
        "transportation": transportation,
        "transport_other": ep.transport_other,
        "traveling_from": ep.traveling_from,
        "returning_to": ep.returning_to,
        "travel_doc_type": travel_doc_type,
        "travel_doc_issue_date": ep.travel_doc_issue_date,  # keep datetime
        "travel_doc_expiry_date": ep.travel_doc_expiry_date,  # keep datetime
        "travel_doc_issued_by": ep.travel_doc_issued_by,
        "bank_name": ep.bank_name,
        "iban": ep.iban,
        "iban_type": iban_type,
        "swift": ep.swift,
    }

    return attendee


# ==============================================================================
# 7. XML Bundle Loader & Lookup Builders
# ==============================================================================

def _load_custom_xml_objects(path: str) -> Optional[Dict[str, Any]]:
    """
    Load structured objects from embedded Custom XML in an Excel file.

    Returns:
        {
            "events": [Event, ...],
            "event":  Event or None,
            "participants": [Participant, ...],
            "participant_events": [EventParticipant, ...]
        }
    """
    records = _collect_custom_xml_records(path)
    if not records:
        return None

    # --- Events ---
    events = []
    for rec in records.get("events", []):
        try:
            events.append(_build_event_from_record(rec))
        except Exception as exc:
            if DEBUG_PRINT:
                print(f"[CUSTOM-XML] Failed to build Event: {exc}")

    # --- Participants ---
    participants: List[Participant] = []
    for rec in records.get("participants", []):
        participant = _build_participant_from_record(rec)
        if participant:
            participants.append(participant)

    # --- Participant ↔ Event relations ---
    participant_events: List[EventParticipant] = []
    for rec in records.get("participant_events", []):
        ep = _build_participant_event_from_record(rec)
        if ep:
            participant_events.append(ep)

    if not events and not participants and not participant_events:
        return None

    return {
        "events": events,
        "event": events[0] if events else None,
        "participants": participants,
        "participant_events": participant_events,
    }

# ==============================================================================
# 8. Lookup Builders (ParticipantsLista / MAIN ONLINE)
# ==============================================================================

def _build_lookup_participantslista(df_positions: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    """
    Build lookup from the 'ParticipantsLista' sheet.

    Key:
        'LAST|First Middle'
    Value:
        {
            "position": ...,
            "phone": ...,
            "email": ...
        }
    """
    name_col  = next((c for c in df_positions.columns if "name (" in c.lower()), None)
    pos_col   = next((c for c in df_positions.columns if "position" in c.lower()), None)
    phone_col = next((c for c in df_positions.columns if "phone" in c.lower()), None)
    email_col = next((c for c in df_positions.columns if "email" in c.lower()), None)

    look: Dict[str, Dict[str, str]] = {}
    if not name_col:
        return look

    for _, row in df_positions.iterrows():
        raw = _normalize(str(row.get(name_col, "")))
        key = _name_key_from_raw(raw)
        if not key:
            continue
        phone_value = normalize_phone(row.get(phone_col, "")) if phone_col else None
        look[key] = {
            "position": _normalize(str(row.get(pos_col, ""))) if pos_col else "",
            "phone":    phone_value or "",
            "email":    _normalize(str(row.get(email_col, ""))) if email_col else "",
        }
    return look


def _build_lookup_main_online(df_online: pd.DataFrame) -> Dict[str, Dict[str, object]]:
    """
    Build lookup from the 'MAIN ONLINE → ParticipantsList' table.

    Key:
        'LAST|First Middle'  (plus fallback 'LAST|First')
    Value:
        normalized field dictionary with translated and enriched values.
    """
    cols = {c.lower().strip(): c for c in df_online.columns}

    def col(label: str) -> Optional[str]:
        return cols.get(label.lower())

    look: Dict[str, Dict[str, object]] = {}
    for _, row in df_online.iterrows():
        first  = _normalize(str(row.get(col("Name")) or ""))
        middle = _normalize(str(row.get(col("Middle name")) or ""))
        last   = _normalize(str(row.get(col("Last name")) or ""))

        if not first and not last:
            continue

        first_middle = " ".join(part for part in [first, middle] if part).strip()
        key  = _name_key(last, first_middle)
        keys = [key]
        if middle and first:
            keys.append(_name_key(last, first))  # Fallback

        # --- Gender normalization ---
        gender_col = col("Gender")
        gender_raw = (str(row.get(gender_col, "")) if gender_col else "").strip()
        normalized_gender = _normalize_gender(gender_raw)
        gender = normalized_gender.value if normalized_gender else gender_raw

        # --- Birth country translation ---
        birth_country_raw  = re.sub(r",\s*world$", "", _normalize(str(row.get(col("Country of Birth"), ""))), flags=re.IGNORECASE)

        # --- Travel document type ---
        travel_doc_type_col = col("Traveling document type")
        travel_doc_type_raw = (
            str(row.get(travel_doc_type_col, "")) if travel_doc_type_col else ""
        ).strip()
        travel_doc_type_value = _normalize_doc_type_label(travel_doc_type_raw)

        # --- Transport and banking fields ---
        transportation_col     = col("Transportation")
        transport_other_col    = col("Transportation (Other)")
        iban_type_col          = col("IBAN Type")

        transportation_value   = str(row.get(transportation_col, "")) if transportation_col else ""
        transport_other_value  = str(row.get(transport_other_col, "")) if transport_other_col else ""
        iban_type_value        = str(row.get(iban_type_col, "")) if iban_type_col else ""

        # --- Compose normalized entry ---
        phone_col = col("Phone number")
        phone_raw = row.get(phone_col, "") if phone_col else ""
        phone_list_value = normalize_phone(phone_raw) or ""

        entry = {
            "name": _to_app_display_name(" ".join([first, middle, last]).strip()),
            "gender": gender,
            "dob": row.get(col("Date of Birth (DOB)")),
            "pob": _normalize(str(row.get(col("Place Of Birth (POB)"), ""))),
            "birth_country": birth_country_raw,
            "citizenships": [
                _normalize(x)
                for x in re.split(r"[;,]", str(row.get(col("Citizenship(s)"), "")))
                if _normalize(x)
            ],
            "email_list": _normalize(str(row.get(col("Email address"), ""))),
            "phone_list": phone_list_value,
            "travel_doc_type": travel_doc_type_value,
            "travel_doc_number": _normalize(str(row.get(col("Traveling document number"), ""))),
            "travel_doc_issue": row.get(col("Traveling document issuance date")),
            "travel_doc_expiry": row.get(col("Traveling document expiration date")),
            "travel_doc_issued_by": translate(
                _normalize(str(row.get(col("Traveling document issued by"), ""))), "en"
            ),
            "transportation_declared": transportation_value.strip(),
            "transport_other": transport_other_value.strip(),
            "traveling_from_declared": _normalize(str(row.get(col("Traveling from"), ""))),
            "returning_to": _normalize(str(row.get(col("Returning to"), ""))),
            "diet_restrictions": _normalize(str(row.get(col("Diet restrictions"), ""))),
            "organization": translate(_normalize(str(row.get(col("Organization"), ""))), "en"),
            "unit": translate(_normalize(str(row.get(col("Unit"), ""))), "en"),
            "rank": translate(_normalize(str(row.get(col("Rank"), ""))), "en"),
            "intl_authority": _normalize(str(row.get(col("Authority"), ""))),
            "bio_short": translate(_normalize(str(row.get(col("Short professional biography"), ""))), "en"),
            "bank_name": _normalize(str(row.get(col("Bank name"), ""))),
            "iban": _normalize(str(row.get(col("IBAN"), ""))),
            "iban_type": iban_type_value.strip(),
            "swift": _normalize(str(row.get(col("SWIFT"), ""))),
        }

        for nk in keys:
            if not nk:
                continue
            if nk not in look:
                look[nk] = entry

    return look


# ==============================================================================
# 9. Column Finder and Main Parsing Routine
# ==============================================================================


def parse_for_commit(path: str, *, preview_only: bool = True) -> dict:
    """
    Parse an Excel workbook into structured event and attendee data.

    Returns:
        dict {
            "event": {...},
            "attendees": [...],
            "objects": None or {events, participants, participant_events},
            "preview": {...}
        }

    Args:
        path: Filesystem path to the uploaded workbook.
        preview_only: Skip participant DB lookups when True (default) to speed
            preview generation.
    """
    # --- Custom XML shortcut (if embedded data exists) ---
    custom_bundle = _load_custom_xml_objects(path)
    if custom_bundle:
        event_obj: Optional[Event] = custom_bundle.get("event")
        participants: List[Participant] = custom_bundle.get("participants", [])
        participant_events: List[EventParticipant] = custom_bundle.get("participant_events", [])

        participants_by_id = {p.pid: p for p in participants}
        attendees = []
        for ep in participant_events:
            participant = participants_by_id.get(ep.participant_id)
            if not participant:
                continue
            attendees.append(_merge_attendee_preview(participant, ep))

        payload = {
            "event": event_obj.model_dump() if event_obj else {},
            "attendees": attendees,
            "objects": custom_bundle,
            "preview": {
                "event": _serialize_event_for_preview(event_obj),
                "participants": [_serialize_participant_for_preview(p) for p in participants],
                "participant_events": [_serialize_participant_event_for_preview(ep) for ep in participant_events],
            },
        }

        if DEBUG_PRINT:
            print("[CUSTOM-XML] Parsed event", payload["preview"]["event"].get("eid"))
            print(f"[CUSTOM-XML] Participants: {len(participants)} | Participant events: {len(participant_events)}")

        return payload

    cache = WorkbookCache(path)
    participant_lookup_enabled = not preview_only

    # --------------------------------------------------------------------------
    # 1. Event Header
    # --------------------------------------------------------------------------
    eid, title, start_date, end_date, place, country, cost = _read_event_header_block(path, cache)
    if DEBUG_PRINT:
        print(
            "[STEP] Event header:",
            {"eid": eid, "title": title, "start_date": start_date,
             "end_date": end_date, "place": place, "country": country, "cost": cost},
        )

    # --------------------------------------------------------------------------
    # 2. Table Discovery & Lookups
    # --------------------------------------------------------------------------
    tables = list_tables(path)
    idx = _index_tables(tables)

    plist = _find_table_exact(idx, "ParticipantsLista")
    ponl = _find_table_exact(idx, "ParticipantsList")

    if not plist:
        raise RuntimeError("Required table 'ParticipantsLista' not found")
    if REQUIRE_PARTICIPANTS_LIST and not ponl:
        raise RuntimeError("Required table 'ParticipantsList' (MAIN ONLINE) not found")

    df_positions = _read_table_df(path, plist, cache)
    df_online = _read_table_df(path, ponl, cache) if ponl else pd.DataFrame()

    positions_lookup = _build_lookup_participantslista(df_positions)
    online_lookup = _build_lookup_main_online(df_online) if not df_online.empty else {}

    if DEBUG_PRINT:
        print(f"[STEP] Positions lookup entries: {len(positions_lookup)}")
        print(f"[STEP] Online lookup entries: {len(online_lookup)}")

    # --------------------------------------------------------------------------
    # 3. Collect Attendees from Country Tables
    # --------------------------------------------------------------------------
    attendees: List[dict] = []
    initial_attendees: List[dict] = []

    for key, country_label in COUNTRY_TABLE_MAP.items():
        table = _find_table_exact(idx, key)
        if not table:
            continue

        df = _read_table_df(path, table, cache)
        if df.empty:
            continue

        # Use the Excel matrix to resolve headers for this country table
        # Matrix shape: {excel_header -> target_field}
        m = get_mapping("Participants", key)  # key is 'tableAlb', 'tableBih', etc.

        # Invert once to target->excel_header for quick lookups
        inv = {t: h for h, t in m.items()}

        # Pick headers from the inverted map. Return None if missing (we'll guard later).
        nm_col = inv.get("name_full")  # was "Name and Last Name"
        trans_col = inv.get("travel")  # was "Travel"
        from_col = inv.get("traveling_from")  # was "Traveling from"
        grade_col = inv.get("grade")  # was "Grade (0 - BL, 1 - Pass, 2 - Excel)"

        # Defensive: if the workbook renamed a header unexpectedly, drop to None.
        if nm_col not in df.columns:
            nm_col = None
        if trans_col not in df.columns:
            trans_col = None
        if from_col not in df.columns:
            from_col = None
        if grade_col not in df.columns:
            grade_col = None

        if not nm_col:
            continue

        prefer_online_transport = trans_col is None

        for _, row in df.iterrows():
            name_cell = row.get(nm_col)
            if name_cell is None or pd.isna(name_cell):
                continue

            if isinstance(name_cell, str) and not name_cell.strip():
                continue

            raw_name = _normalize(str(name_cell))
            if not raw_name or raw_name.upper() == "TOTAL":
                continue

            def _normalized_cell(col_name: Optional[str]) -> str:
                if not col_name:
                    return ""
                value = row.get(col_name)
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    return ""
                return _normalize(str(value))

            transportation = _normalized_cell(trans_col)
            traveling_from = _normalized_cell(from_col)
            grade_val = row.get(grade_col, None)
            grade = None
            if isinstance(grade_val, (int, float)) and not pd.isna(grade_val):
                try:
                    grade = int(grade_val)
                except Exception:
                    pass

            # --- Match lookups ---
            variants = list(_split_name_variants(raw_name))
            p_list, p_comp = {}, {}
            for f, m, l in variants:
                key_a = _name_key(l, " ".join([f, m]).strip())
                key_b = _name_key(l, f) if f else None
                cand_list = (online_lookup.get(key_a) or (online_lookup.get(key_b) if key_b else None)) or {}
                cand_comp = (positions_lookup.get(key_a) or (positions_lookup.get(key_b) if key_b else None)) or {}
                if cand_list or cand_comp:
                    p_list, p_comp = cand_list, cand_comp
                    break

            # --- Base attendee record ---
            ordered = _normalize(raw_name)
            if "," in ordered:
                last_part, first_part = [x.strip() for x in ordered.split(",", 1)]
                ordered = f"{first_part} {last_part}".strip()
            base_name = _to_app_display_name(ordered)

            country_cid = get_country_cid_by_name(country_label) or country_label
            if transportation:
                transportation_value = transportation
            elif prefer_online_transport:
                transportation_value = p_list.get("transportation_declared") or ""
            else:
                transportation_value = ""
            transport_other_value = (str(p_list.get("transport_other", "")) or "").strip()
            traveling_from_value = traveling_from or p_list.get("traveling_from_declared") or ""
            grade_value = grade if grade is not None else int(Grade.NORMAL)

            base_record = {
                "name": base_name,
                "representing_country": country_cid,
                "transportation": transportation_value,
                "transport_other": transport_other_value,
                "traveling_from": traveling_from_value,
                "grade": grade_value,
            }
            initial_attendees.append(base_record)

            # --- Enrich with ParticipantsLista info ---
            record = {
                **base_record,
                "position": p_comp.get("position") or "",
                "phone": normalize_phone(p_comp.get("phone")) or "",
                "email": p_comp.get("email") or "",
            }

            # --- MAIN ONLINE enrichment ---
            online = p_list or {}
            _fill_if_missing(record, "position", online, "position_online")
            _fill_if_missing(record, "phone", online, "phone_list")
            _fill_if_missing(record, "email", online, "email_list")
            _fill_if_missing(record, "traveling_from", online, "traveling_from_declared")
            if not record.get("transportation") and prefer_online_transport:
                record["transportation"] = online.get("transportation_declared") or None
            _fill_if_missing(record, "transport_other", online, "transport_other")

            # --- Country & citizenship normalization ---
            birth_country_value = online.get("birth_country", "")
            birth_res = resolve_country_flexible(str(birth_country_value))
            birth_country_cid = birth_res["cid"] if birth_res else country_cid
            citizenships_raw = online.get("citizenships", [])
            if isinstance(citizenships_raw, str):
                citizenships_raw = [citizenships_raw]

            citizenships_clean: list[str] = []
            for tok in _split_multi_country(citizenships_raw):
                res = resolve_country_flexible(tok)
                if res and res.get("cid"):
                    cid = res["cid"]
                    if cid not in citizenships_clean:
                        citizenships_clean.append(cid)

            if DEBUG_PRINT:
                print("[TOKENS]", _split_multi_country(online.get("citizenships", [])))
            for tok in _split_multi_country(online.get("citizenships", [])):
                r = resolve_country_flexible(tok)
                print("   ->", tok, "=>", (r and r.get("cid"), r and r.get("country")))
            if DEBUG_PRINT:
                print("[OUT] citizenships:", citizenships_clean)


            # --- Final enrichment ---
            record.update({
                "gender": online.get("gender", ""),
                "dob": date_to_iso(online.get("dob"), tzinfo=EU_TZ),
                "pob": online.get("pob", ""),
                "birth_country": birth_country_cid,
                "citizenships": citizenships_clean,
                "travel_doc_type": online.get("travel_doc_type"),
                "travel_doc_number": online.get("travel_doc_number", ""),
                "travel_doc_issue_date": date_to_iso(online.get("travel_doc_issue"), tzinfo=EU_TZ),
                "travel_doc_expiry_date": date_to_iso(online.get("travel_doc_expiry"), tzinfo=EU_TZ),
                "travel_doc_issued_by": online.get("travel_doc_issued_by", ""),
                "returning_to": online.get("returning_to", ""),
                "diet_restrictions": online.get("diet_restrictions", ""),
                "organization": online.get("organization", ""),
                "unit": online.get("unit", ""),
                "rank": online.get("rank", ""),
                "intl_authority": _parse_bool_value(online.get("intl_authority", "")) or False,
                "bio_short": online.get("bio_short", ""),
                "bank_name": online.get("bank_name", ""),
                "iban": online.get("iban", ""),
                "iban_type": online.get("iban_type"),
                "swift": online.get("swift", ""),
            })
            if DEBUG_PRINT:
                print(f"[DEBUG] citizenships_in={online.get('citizenships')} → {record['citizenships']}")

            participant = None
            if participant_lookup_enabled:
                participant = lookup(
                    name_display=record.get("name", ""),
                    country_name=country_label,
                    dob_source=online.get("dob"),
                    representing_country=country_cid,
                )

            if participant:
                record["pid"] = participant.pid

            record["phone"] = normalize_phone(record.get("phone")) or ""
            attendees.append(record)

    # --------------------------------------------------------------------------
    # 4. Assemble Final Payload
    # --------------------------------------------------------------------------
    if DEBUG_PRINT:
        print("[STEP] Initial participant list:")
        for rec in initial_attendees:
            print("  ", rec)

    payload = {
        "event": {
            "eid": eid,
            "title": title,
            "start_date": start_date,
            "end_date": end_date,
            "place": place,
            "country": country,
            "type": "Training",
            "cost": cost,
        },
        "attendees": attendees,
        "objects": None,
        "preview": {
            "event": {
                "eid": eid,
                "title": title,
                "start_date": date_to_iso(start_date, tzinfo=EU_TZ),
                "end_date": date_to_iso(end_date, tzinfo=EU_TZ),
                "place": place,
                "country": country,
                "type": "Training",
                "cost": cost,
            },
            "participants": attendees,
            "participant_events": [],
        },
    }

    if DEBUG_PRINT:
        payload["initial_attendees"] = initial_attendees

    return payload

# ==============================================================================
# 10. String / Normalization Helpers
# ==============================================================================

def _normalize(s: Optional[str]) -> str:
    """Normalize whitespace and coerce None to an empty string."""
    return re.sub(r"\s+", " ", (s or "").strip())


def _normalize_doc_type_label(value: object) -> str:
    """Return 'Passport' only if value == 'Passport'; everything else 'ID Card'."""
    if DEBUG_PRINT:
        print(f"[DEBUG] Normalizing doc type label: {value}")

    if value == "Passport":
        return str(DocType.passport.value)
    return str(DocType.id_card.value)
    # if value is None:
    #     return ""
    #
    # text = str(value).strip()
    # if not text:
    #     return ""
    #
    # slug = re.sub(r"[^a-z0-9]+", "", text.lower())
    # passport_slug = re.sub(r"[^a-z0-9]+", "", DocType.passport.value.lower())
    # if slug == passport_slug:
    #     return str(DocType.passport.value)
    # return str(DocType.id_card.value)


# ==============================================================================
# 11. Workbook / Table Utilities
# ==============================================================================

def _index_tables(tables: List[TableRef]) -> Dict[str, List[TableRef]]:
    """Group tables by normalized name for quick lookup."""
    idx: Dict[str, List[TableRef]] = {}
    for t in tables:
        idx.setdefault(t.name_norm, []).append(t)
    return idx


def _find_table_exact(idx: Dict[str, List[TableRef]], desired: str) -> Optional[TableRef]:
    """Find the first table matching `desired` by normalized name."""
    target = _norm_tablename(desired)
    group = idx.get(target)
    return group[0] if group else None


def _read_table_df(path: str, table: TableRef, cache: WorkbookCache | None = None) -> pd.DataFrame:
    """
    Read a ListObject range (e.g. 'A4:K7') into a DataFrame.
    Uses the header row as columns.
    """
    def _build_df(ws) -> pd.DataFrame:
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        rows = list(
            ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )
        return _dataframe_from_rows(rows)

    if cache:
        return cache.get_table_df(table, _build_df)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[table.sheet_title]
    return _build_df(ws)


def _dataframe_from_rows(rows: List[tuple]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    header = [_normalize(str(h)) if h is not None else "" for h in rows[0]]
    df = pd.DataFrame(rows[1:], columns=header).dropna(how="all")

    empty_cols = [col for col in df.columns if not str(col).strip()]
    if empty_cols:
        df = df.drop(columns=empty_cols)

    return df


# ==============================================================================
# 12. Event-Header Parsing
# ==============================================================================

def _filename_year_from_eid(filename: str) -> int:
    """Infer 4-digit year from file name pattern like 'PFE25M2' → 2025."""
    m = re.search(r"PFE(\d{2})M", filename.upper())
    return 2000 + int(m.group(1)) if m else datetime.now(UTC).year


def _parse_event_header(a1: str, a2: str, year: int):
    """
    Parse event metadata from:
      A1 → 'PFE25M2 TITLE OF EVENT'
      A2 → 'JUNE 23 - 27 - Opatija, CROATIA'
    """
    a1 = _normalize(a1)
    sp = a1.find(" ")
    eid, title = (a1, "") if sp == -1 else (a1[:sp], a1[sp + 1:])

    a2 = _normalize(a2)
    parts = [p.strip() for p in a2.split(" - ")]
    start_date = end_date = None
    location = ""

    if len(parts) >= 3:
        month_and_start, end_day_str, location = parts[0], parts[1], parts[2]
        m = re.match(r"([A-Z]+)\s+(\d{1,2})", month_and_start.upper())
        if m:
            month_num = MONTHS.get(m.group(1))
            start_day = int(m.group(2))
            if month_num:
                end_day = int(re.sub(r"\D", "", end_day_str))
                start_date = datetime(year, month_num, start_day, tzinfo=UTC)
                end_date = datetime(year, month_num, end_day, tzinfo=UTC)

    place = location
    country_value: str | None = None
    if location:
        loc_parts = [p.strip() for p in location.split(",", 1)]
        place = loc_parts[0]
        raw_country = loc_parts[1] if len(loc_parts) > 1 else ""
        if raw_country:
            normalized_country = _normalize(raw_country)
            lookup = get_country_cid_by_name(normalized_country) or get_country_cid_by_name(normalized_country.title())
            country_value = lookup or normalized_country

    return eid, title, start_date, end_date, place, country_value


def _read_event_header_block(
    path: str,
    cache: WorkbookCache | None = None,
) -> tuple[str, str, datetime, datetime, str, Optional[str], Optional[float]]:
    """Read event header data from the Participants and COST Overview sheets."""
    wb = cache.get_workbook() if cache else openpyxl.load_workbook(path, data_only=True)
    if "Participants" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Participants' not found")
    if "COST Overview" not in wb.sheetnames:
        raise RuntimeError("Sheet 'COST Overview' not found")

    ws = wb["Participants"]
    a1 = ws["A1"].value or ""
    a2 = ws["A2"].value or ""
    year = _filename_year_from_eid(os.path.basename(path))
    eid, title, start_date, end_date, place, country = _parse_event_header(a1, a2, year)

    wws = wb["COST Overview"]
    cost_overview_b15 = str(wws["B15"].value or "").strip()
    cost = float(cost_overview_b15) if cost_overview_b15 else None
    return eid, title, start_date, end_date, place, country, cost


# ==============================================================================
# 13. Public Validation / Preview API
# ==============================================================================

def validate_excel_file_for_import(path: str) -> tuple[bool, list[str], dict]:
    """
    Validate Excel workbook structure and required elements.
    Checks:
      - Sheets: Participants, COST Overview
      - Cells: A1/A2, B15
      - Tables: ParticipantsLista, ≥1 country table
    """
    custom_bundle = _load_custom_xml_objects(path)
    if custom_bundle:
        event_obj = custom_bundle.get("event")
        participants = custom_bundle.get("participants", [])
        participant_events = custom_bundle.get("participant_events", [])
        seen = {
            "custom_xml": True,
            "events": len(custom_bundle.get("events", [])),
            "participants": len(participants),
            "participant_events": len(participant_events),
        }
        if event_obj:
            seen["event_eid"] = event_obj.eid
        return True, [], seen

    missing: list[str] = []

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Participants" not in wb.sheetnames:
        missing.append("Sheet 'Participants'")
        return False, missing, {}
    if "COST Overview" not in wb.sheetnames:
        missing.append("Sheet 'COST Overview'")
        return False, missing, {}

    ws, wws = wb["Participants"], wb["COST Overview"]
    a1 = (ws["A1"].value or "").strip()
    a2 = (ws["A2"].value or "").strip()
    cost_overview_b15 = str(wws["B15"].value or "").strip()
    if not a1:
        missing.append("Participants!A1 (eid + title)")
    if not a2:
        missing.append("Participants!A2 (dates + location)")
    if not cost_overview_b15:
        missing.append("Cost Overview!B15 (Total Cost)")

    tables = list_tables(path)
    idx = _index_tables(tables)

    if not _find_table_exact(idx, "ParticipantsLista"):
        missing.append("Table 'ParticipantsLista'")
    if not any(_find_table_exact(idx, k) for k in COUNTRY_TABLE_MAP.keys()):
        missing.append("At least one country table (tableAlb, tableBih, tableCro, etc.)")

    ok = len(missing) == 0
    seen = {t.name_norm: (t.name, t.sheet_title, t.ref) for t in tables}

    if DEBUG_PRINT:
        print("[VALIDATE] OK:", ok)
        if missing:
            print("[VALIDATE] Missing:", missing)
        print("[VALIDATE] Tables seen:", seen)

    return ok, missing, seen


def inspect_and_preview_uploaded(path: str, *, preview_only: bool = True) -> None:
    """
    Dry-run inspection: prints existing/new event info and attendee summaries.
    No DB writes.

    Args:
        path: Filesystem path to the uploaded workbook.
        preview_only: Skip participant DB lookups when True (default).
    """
    cache = WorkbookCache(path)
    participant_lookup_enabled = not preview_only
    eid, title, start_date, end_date, place, country, cost = _read_event_header_block(path, cache)

    existing = mongodb.collection("events").find_one({"eid": eid})
    if existing:
        print(f"[EVENT] EXIST {eid} title='{existing.get('title','')}' "
              f"start_date={existing.get('start_date')} place='{existing.get('place','')}' "
              f"country='{existing.get('country')}'")
    else:
        print(f"[EVENT] NEW {eid} title='{title}' start_date={start_date} "
              f"end_date={end_date} place='{place}' country='{country}'")

    tables = list_tables(path)
    idx = _index_tables(tables)

    plist = _find_table_exact(idx, "ParticipantsLista")
    if not plist:
        raise RuntimeError("Required table 'ParticipantsLista' not found (any sheet)")

    df_positions = _read_table_df(path, plist, cache)
    positions_lookup_full = _build_lookup_participantslista(df_positions)

    print("[ATTENDEES]")
    for key, country_label in COUNTRY_TABLE_MAP.items():
        t = _find_table_exact(idx, key)
        if not t:
            continue
        df = _read_table_df(path, t, cache)
        if df.empty:
            continue

        nm_col = next((c for c in df.columns if "name" in c.lower()), None)
        grade_col = next((c for c in df.columns if "grade" in c.lower()), None)

        if not nm_col:
            continue

        for _, row in df.iterrows():
            name_cell = row.get(nm_col)
            if name_cell is None or pd.isna(name_cell):
                continue

            if isinstance(name_cell, str) and not name_cell.strip():
                continue

            raw_name = _normalize(str(name_cell))
            if not raw_name:
                continue
            grade = _normalize(str(row.get(grade_col, ""))) if grade_col else ""
            key_lookup = _name_key_from_raw(raw_name)
            pos = positions_lookup_full.get(key_lookup, {}).get("position", "")

            participant = None
            if participant_lookup_enabled:
                participant = lookup(
                    name_display=raw_name,
                    country_name=country_label,
                )

            norm = _to_app_display_name(raw_name)
            star = "*" if not participant else " "
            pid = participant.pid if participant else "NEW"

            print(f"{star} {'NEW' if star=='*' else 'EXIST'} {pid:>6} {norm} "
                  f"({grade}, {country_label}) {'pos='+pos if pos else ''}")


# ==============================================================================
# 15. Utility Fill Helper
# ==============================================================================

def _name_key_from_raw(raw_display: str) -> str:
    """Normalize 'Last, First' or 'First Last' → 'last|first' key."""
    s = _normalize(raw_display)
    if not s:
        return ""
    if "," in s:
        last, first = [x.strip() for x in s.split(",", 1)]
    else:
        parts = s.split()
        last = parts[-1] if len(parts) > 1 else s
        first = " ".join(parts[:-1]) if len(parts) > 1 else ""
    return _name_key(last, first)


def _fill_if_missing(dst: dict, key: str, src: dict, src_key: Optional[str] = None) -> None:
    """If dst[key] empty, copy src[src_key or key] if truthy."""
    k = src_key or key
    if dst.get(key):
        return

    value = src.get(k)
    if not value:
        return

    if key == "phone":
        normalized = normalize_phone(value)
        if not normalized:
            return
        dst[key] = normalized
        return

    dst[key] = value
