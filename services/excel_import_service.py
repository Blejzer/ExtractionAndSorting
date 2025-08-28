# services/excel_import_service.py
from __future__ import annotations

import os
import re
import unicodedata
from typing import Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
from datetime import datetime, UTC

from config.database import mongodb_connection
from services.xlsx_tables_inspector import (
    list_sheets,
    list_tables,
    TableRef,
)

# ============================
# Configuration / Constants
# ============================

COUNTRY_TABLE_MAP: Dict[str, str] = {
    "tableAlb": "Albania",
    "tableBih": "Bosnia and Herzegovina",
    "tableCro": "Croatia",
    "tableKos": "Kosovo",
    "tableMne": "Montenegro",
    "tableNmk": "North Macedonia",
    "tableSer": "Serbia",
}

MONTHS: Dict[str, int] = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4, "MAY": 5, "JUNE": 6,
    "JULY": 7, "AUGUST": 8, "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12
}

DEBUG_PRINT = True  # flip to False to quiet logs after you’re happy


# ============================
# String / Name helpers
# ============================

def _normalize(s: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def _strip_accents(text: str) -> str:
    if not text:
        return ""
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))

def _norm_tablename(name: str) -> str:
    return re.sub(r"[^0-9a-zA-Z]+", "", (name or "")).lower()

def _to_app_display_name(fullname: str) -> str:
    """'First Middle Last' -> 'First Middle LAST' (app convention)."""
    name = _normalize(fullname)
    parts = name.split(" ")
    if len(parts) <= 1:
        return name
    return " ".join(parts[:-1]) + " " + parts[-1].upper()


# ============================
# Workbook / Table helpers
# ============================

def _index_tables(tables: List[TableRef]) -> Dict[str, List[TableRef]]:
    """Group tables by normalized name for quick lookup."""
    idx: Dict[str, List[TableRef]] = {}
    for t in tables:
        idx.setdefault(t.name_norm, []).append(t)
    return idx

def _find_table_any(idx: Dict[str, List[TableRef]], desired: str) -> Optional[TableRef]:
    """
    Find a table whose normalized name matches exactly, or
    whose normalized name starts with the desired normalized value.
    """
    target = _norm_tablename(desired)
    # exact
    if target in idx and idx[target]:
        return idx[target][0]
    # prefix
    for key, group in idx.items():
        if key.startswith(target) and group:
            return group[0]
    return None

def _find_table_exact(idx: dict[str, list[TableRef]], desired: str) -> Optional[TableRef]:
    """Match ONLY the exact normalized table name."""
    target = _norm_tablename(desired)
    group = idx.get(target)
    return group[0] if group else None

def _read_table_df(path: str, table: TableRef) -> pd.DataFrame:
    """
    Read a ListObject range (e.g., 'A4:K7') from the given sheet into a DataFrame.
    Uses the header row in the table as columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[table.sheet_title]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col, values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [_normalize(str(h)) if h is not None else "" for h in rows[0]]
    return pd.DataFrame(rows[1:], columns=header).dropna(how="all")


# ============================
# Event header parsing (A1, A2)
# ============================

def _filename_year_from_eid(filename: str) -> int:
    r"""
    From 'PFE25M2 - whatever.xlsx' infer 2025 via regex: r'PFE(\d{2})M'.
    Fallback to current UTC year if not found.
    """
    m = re.search(r"PFE(\d{2})M", filename.upper())
    return 2000 + int(m.group(1)) if m else datetime.now(UTC).year

def _parse_event_header(a1: str, a2: str, year: int):
    """
    A1: 'PFE25M2 TITLE OF THE EVENT' -> (eid, title)
    A2: 'JUNE 23 - 27 - Opatija, CROATIA' -> (dateFrom, dateTo, location)
    Month for end date assumed same as start (per your spec).
    """
    a1 = _normalize(a1)
    sp = a1.find(" ")
    eid, title = (a1, "") if sp == -1 else (a1[:sp], a1[sp + 1:])

    a2 = _normalize(a2)
    parts = [p.strip() for p in a2.split(" - ")]
    date_from = date_to = None
    location = ""
    if len(parts) >= 3:
        month_and_start, end_day_str, location = parts[0], parts[1], parts[2]
        m = re.match(r"([A-Z]+)\s+(\d{1,2})", month_and_start.upper())
        if m:
            month_num = MONTHS.get(m.group(1))
            start_day = int(m.group(2))
            if month_num:
                end_day = int(re.sub(r"\D", "", end_day_str))
                date_from = datetime(year, month_num, start_day, tzinfo=UTC)
                date_to = datetime(year, month_num, end_day, tzinfo=UTC)
    return eid, title, date_from, date_to, location


# ============================
# DB helpers (read-only)
# ============================

def _country_id(country: str):
    doc = mongodb_connection.countries.find_one({"country": country})
    return doc["_id"] if doc else None

def _participant_exists(name_display: str, country_name: str):
    """
    Existence check by normalized app name (First ... LAST) + country_id if available.
    """
    q = {"name": _to_app_display_name(name_display)}
    cid = _country_id(country_name)
    if cid is not None:
        q["country_id"] = cid
    doc = mongodb_connection.participants.find_one(q)
    return (doc is not None), (doc or {})


# ============================
# Public API
# ============================

def validate_excel_file_for_import(path: str) -> tuple[bool, list[str], dict]:
    """
    Validate that:
      - Sheet 'Participants' exists
      - A1 (eid+title) and A2 (dates+location) present (non-empty)
      - Table 'ParticipantsLista' exists (any sheet)
      - At least one of the country tables exists (any sheet)
    Returns: (ok, missing_list, tables_info_dict)
    """
    missing: list[str] = []

    # 1) A1/A2 on "Participants"
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Participants" not in wb.sheetnames:
        missing.append("Sheet 'Participants'")
        return False, missing, {}
    ws = wb["Participants"]
    a1 = (ws["A1"].value or "").strip()
    a2 = (ws["A2"].value or "").strip()
    if not a1:
        missing.append("Participants!A1 (eid + title)")
    if not a2:
        missing.append("Participants!A2 (dates + location)")

    # 2) Tables via inspector (cross-sheet, ZIP-safe)
    tables = list_tables(path)
    idx = _index_tables(tables)

    # ParticipantsLista present?
    if not _find_table_exact(idx, "ParticipantsLista"):
        missing.append("Table 'ParticipantsLista'")

    # At least one country table present?
    if not any(_find_table_exact(idx, k) for k in COUNTRY_TABLE_MAP.keys()):
        missing.append("At least one country table (tableAlb, tableBih, tableCro, tableKos, tableMne, tableNmk, tableSer)")

    ok = len(missing) == 0

    # For visibility return a compact dict of what we saw
    seen = {t.name_norm: (t.name, t.sheet_title, t.ref) for t in tables}

    if DEBUG_PRINT:
        print("[VALIDATE] OK:", ok)
        if missing:
            print("[VALIDATE] Missing:", missing)
        print("[VALIDATE] Tables seen:", seen)

    return ok, missing, seen


def inspect_and_preview_uploaded(path: str) -> None:
    """
    NO WRITES.
    - Logs EVENT NEW/EXIST (from A1/A2 on 'Participants' sheet)
    - Logs ATTENDEES from any present country tables
      (marks new with '*', includes grade, country, and position from ParticipantsLista)
    """
    # Event header
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Participants" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Participants' not found")
    ws = wb["Participants"]
    a1 = ws["A1"].value or ""
    a2 = ws["A2"].value or ""
    year = _filename_year_from_eid(os.path.basename(path))
    eid, title, date_from, date_to, location = _parse_event_header(a1, a2, year)

    # Event exist check (read-only)
    existing = mongodb_connection.events.find_one({"eid": eid})
    if existing:
        print(f"[EVENT] EXIST {eid}  title='{existing.get('title','')}' "
              f"dateFrom={existing.get('dateFrom')} location='{existing.get('location','')}'")
    else:
        print(f"[EVENT] NEW   {eid}  title='{title}' dateFrom={date_from} dateTo={date_to} location='{location}'")

    # Tables + positions
    tables = list_tables(path)
    idx = _index_tables(tables)

    plist = _find_table_exact(idx, "ParticipantsLista")
    if not plist:
        raise RuntimeError("Required table 'ParticipantsLista' not found (any sheet)")

    df_positions = _read_table_df(path, plist)
    name_col_pos = next((c for c in df_positions.columns if "name (" in c.lower()), None)
    pos_col = next((c for c in df_positions.columns if "position" in c.lower()), None)

    positions_lookup: Dict[str, str] = {}
    if name_col_pos and pos_col:
        for _, r in df_positions.iterrows():
            raw = _normalize(str(r.get(name_col_pos, "")))
            pos = _normalize(str(r.get(pos_col, "")))
            if not raw:
                continue
            # raw is "LAST, First Middle"
            if "," in raw:
                last, first = [x.strip() for x in raw.split(",", 1)]
            else:
                parts = raw.split(" ")
                last = parts[-1] if len(parts) > 1 else raw
                first = " ".join(parts[:-1])
            key = (_strip_accents(last) + "|" + _strip_accents(first)).lower()
            positions_lookup[key] = pos

    print("[ATTENDEES]")

    # Iterate any present country tables
    for key, country_label in COUNTRY_TABLE_MAP.items():
        t = _find_table_exact(idx, key)
        if not t:
            continue

        df = _read_table_df(path, t)
        if df.empty:
            continue

        nm_col = next((c for c in df.columns if "name" in c.lower()), None)
        grade_col = next((c for c in df.columns if "grade" in c.lower()), None)

        for _, row in df.iterrows():
            raw_name = _normalize(str(row.get(nm_col, ""))) if nm_col else ""
            if not raw_name:
                continue
            grade = _normalize(str(row.get(grade_col, ""))) if grade_col else ""

            # Build key to lookup position: "LAST|First Middle"
            parts = raw_name.split(" ")
            if len(parts) > 1:
                first = " ".join(parts[:-1])
                last = parts[-1]
            else:
                first = ""
                last = parts[0]
            key_lookup = (_strip_accents(last) + "|" + _strip_accents(first)).lower()
            pos = positions_lookup.get(key_lookup, "")

            exists, doc = _participant_exists(raw_name, country_label)
            norm = _to_app_display_name(raw_name)
            star = "*" if not exists else " "
            pid = doc.get("pid", "NEW")

            print(
                f"{star} {'NEW' if star=='*' else 'EXIST'} {pid:>6}  {norm}  ({grade}, {country_label})  "
                f"{'pos='+pos if pos else ''}"
            )
