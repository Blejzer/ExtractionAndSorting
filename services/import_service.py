# services/import_service.py
"""Excel import/validation helpers.

This service locates Excel Tables (ListObjects) across all worksheets and performs:
- Structural validation (required table names must be present)
- Optional preview: event header extraction (Participants!A1/A2), discovered tables, etc.
- Dry-run parsing that prints diagnostic output (no DB writes here)

Conventions
-----------
- Country table names are looked up exactly (e.g., 'tableAlb', 'tableBih', etc.).
- 'ParticipantsLista' must exist for position lookups.
- Titles and dates are read from the 'Participants' worksheet in cells A1 and A2.
"""

from __future__ import annotations

import os
import re
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime, UTC, time
from typing import Dict, List, Optional, Iterator, Any, Mapping

import openpyxl
import pandas as pd
from openpyxl.utils import range_boundaries

from config.database import mongodb
from domain.models.event import Event, EventType
from domain.models.event_participant import (
    DocType,
    EventParticipant,
    IbanType,
    Transport,
)
from domain.models.participant import Grade, Gender, Participant

from repositories.event_repository import EventRepository
from repositories.participant_event_repository import ParticipantEventRepository
from repositories.participant_repository import ParticipantRepository

from services.xlsx_tables_inspector import (
    list_tables,
    TableRef,
)
from utils.country_resolver import (
    COUNTRY_TABLE_MAP,
    normalize_citizenships,
    resolve_birth_country_cid,
)
from utils.dates import MONTHS
from utils.helpers import _normalize_gender
from utils.translation import translate

# ============================
# Configuration / Constants
# ============================

DEBUG_PRINT = False  # flip to True for verbose logging and extra debug data

# --- ADD near the top with other constants ---
# We now require the full roster table for enrichment:
REQUIRE_PARTICIPANTS_LIST = True


# --- ADD under existing helpers: name/build-key utilities ---
def _canon(name: str) -> str:
    """Return a lowercase, accent-stripped version of ``name``."""
    if not name:
        return ""
    nfd = unicodedata.normalize("NFD", name)
    return "".join(ch for ch in nfd if not unicodedata.combining(ch)).lower()


def _name_key(last: str, first_middle: str) -> str:
    return f"{_canon(last)}|{_canon(first_middle)}".strip()

def _split_name_variants(raw: str) -> Iterator[tuple[str, str, str]]:
    """Yield (first, middle, last) variants for a raw name string.

    The last 1-3 tokens are treated as possible surnames. Names may also be
    provided as ``LAST, First Middle``; in that case the tokens are reordered
    to ``First Middle LAST`` before generating variants.
    """
    s = _normalize(raw)
    if not s:
        return

    if "," in s:
        last_part, first_part = [x.strip() for x in s.split(",", 1)]
        tokens = first_part.split() + last_part.split()
    else:
        tokens = s.split()

    tokens = [_canon(t) for t in tokens]

    if len(tokens) == 1:
        yield tokens[0], "", ""
        return

    max_surname = min(3, len(tokens) - 1)
    for i in range(1, max_surname + 1):
        first_middle = tokens[:-i]
        last_tokens = tokens[-i:]
        first = first_middle[0]
        middle = " ".join(first_middle[1:]) if len(first_middle) > 1 else ""
        last = " ".join(last_tokens)
        yield first, middle, last


# ============================
# Custom XML extraction helpers
# ============================

def _strip_xml_tag(tag: str) -> str:
    return tag.split("}", 1)[1] if "}" in tag else tag


def _element_to_flat_dict(elem: ET.Element, prefix: str = "") -> Dict[str, str]:
    children = list(elem)
    if not children:
        key = prefix or _strip_xml_tag(elem.tag)
        return {key: (elem.text or "").strip()}

    data: Dict[str, str] = {}
    for child in children:
        child_tag = _strip_xml_tag(child.tag)
        child_prefix = f"{prefix}_{child_tag}" if prefix else child_tag
        child_data = _element_to_flat_dict(child, child_prefix)
        for key, value in child_data.items():
            if not value:
                continue
            if key in data and data[key]:
                data[key] = f"{data[key]}; {value}"
            else:
                data[key] = value
    return data


def _collect_custom_xml_records(path: str) -> Optional[Dict[str, List[Dict[str, str]]]]:
    try:
        with zipfile.ZipFile(path) as zf:
            names = [
                name
                for name in zf.namelist()
                if name.startswith("customXml/") and name.endswith(".xml")
            ]
            if not names:
                return None

            collected: Dict[str, List[Dict[str, str]]] = {
                "participant": [],
                "event": [],
                "participant_event": [],
            }

            for name in names:
                try:
                    root = ET.fromstring(zf.read(name))
                except ET.ParseError:
                    if DEBUG_PRINT:
                        print(f"[CUSTOM-XML] Failed to parse {name}")
                    continue

                stack = [root]
                while stack:
                    node = stack.pop()
                    tag = _strip_xml_tag(node.tag)
                    if tag in collected:
                        collected[tag].append(_element_to_flat_dict(node))
                    stack.extend(list(node))

            if not any(collected.values()):
                return None

            return {
                "participants": collected["participant"],
                "events": collected["event"],
                "participant_events": collected["participant_event"],
            }
    except (zipfile.BadZipFile, FileNotFoundError):
        return None

def _parse_bool_value(value: object) -> Optional[bool]:
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s:
        return None
    if s in {"true", "1", "yes"}:
        return True
    if s in {"false", "0", "no"}:
        return False
    return None


def _coerce_event_type(value: object) -> Optional[EventType]:
    if value is None:
        return None
    candidates = []
    if isinstance(value, EventType):
        return value
    s = str(value).strip()
    if not s:
        return None
    candidates.append(s)
    candidates.append(s.title())
    candidates.append(s.upper())
    for cand in candidates:
        try:
            return EventType(cand)
        except ValueError:
            continue
    return None


def _coerce_grade_value(value: object) -> int:
    if isinstance(value, (int, float)) and not pd.isna(value):
        return int(value)
    if value is None:
        return int(Grade.NORMAL)
    s = str(value).strip()
    if not s:
        return int(Grade.NORMAL)
    try:
        return int(float(s))
    except ValueError:
        pass
    try:
        return int(Grade[s.upper()].value)
    except Exception:
        return int(Grade.NORMAL)

def _parse_datetime_value(value: object) -> Optional[datetime]:
    """Coerce Excel/Pandas/strings/date into timezone-aware datetime (UTC, 00:00)."""
    if isinstance(value, datetime):
        # if naive, assume UTC
        return value if value.tzinfo else value.replace(tzinfo=UTC)
    # Excel/Pandas might pass numpy datetime64 or pandas Timestamp
    try:
        import pandas as _pd  # local import to avoid hard dep at module import
        if isinstance(value, _pd.Timestamp):
            dt = value.to_pydatetime()
            return dt if dt.tzinfo else dt.replace(tzinfo=UTC)
    except Exception:
        pass

    # If it's a date, upcast to datetime @ 00:00 UTC
    try:
        # duck-typing to avoid importing date
        if value.__class__.__name__ == "date":
            return datetime(value.year, value.month, value.day, tzinfo=UTC)
    except Exception:
        pass

    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    # ISO or common formats
    for try_fn in (
        lambda x: datetime.fromisoformat(x),
        lambda x: datetime.strptime(x, "%Y-%m-%d"),
        lambda x: datetime.strptime(x, "%d.%m.%Y"),
        lambda x: datetime.strptime(x, "%m/%d/%Y"),
    ):
        try:
            dt = try_fn(s)
            return dt if dt.tzinfo else dt.replace(tzinfo=UTC)
        except Exception:
            continue

    return None



def _build_event_from_record(record: Dict[str, str]) -> Event:
    start_date = _parse_datetime_value(record.get("start_date"))
    end_date = _parse_datetime_value(record.get("end_date"))
    cost_raw = record.get("cost")
    cost_val: Optional[float] = None
    if cost_raw not in (None, ""):
        try:
            cost_val = float(str(cost_raw).strip())
        except ValueError:
            cost_val = None

    event_type = _coerce_event_type(record.get("type"))

    return Event(
        eid=str(record.get("eid", "")),
        title=str(record.get("title", "")),
        start_date=start_date,
        end_date=end_date,
        place=str(record.get("place", "")),
        country=record.get("country") or None,
        type=event_type,
        cost=cost_val,
    )


def _build_participant_from_record(record: Dict[str, str]) -> Optional[Participant]:
    data: Dict[str, Any] = dict(record)
    normalized_gender = _normalize_gender(data.get("gender"))
    if normalized_gender is not None:
        data["gender"] = normalized_gender
    data["grade"] = _coerce_grade_value(data.get("grade"))
    dob_dt = _parse_datetime_value(data.get("dob"))
    if dob_dt:
        data["dob"] = dob_dt
    bool_fields = ["intl_authority"]
    for field in bool_fields:
        if field in data:
            val = _parse_bool_value(data[field])
            if val is not None:
                data[field] = val
    try:
        return Participant.model_validate(data)
    except Exception as exc:
        if DEBUG_PRINT:
            print(f"[CUSTOM-XML] Failed to build Participant: {exc}")
        return None


def _build_participant_event_from_record(record: Dict[str, str]) -> Optional[EventParticipant]:
    data: Dict[str, Any] = dict(record)
    for key in ("travel_doc_issue_date", "travel_doc_expiry_date"):
        if key in data:
            data[key] = _parse_datetime_value(data.get(key))
    try:
        return EventParticipant.model_validate(data)
    except Exception as exc:
        if DEBUG_PRINT:
            print(f"[CUSTOM-XML] Failed to build EventParticipant: {exc}")
        return None


def _serialize_event_for_preview(event: Optional[Event]) -> Dict[str, Any]:
    if not event:
        return {}
    data = event.model_dump()
    preview: Dict[str, Any] = {}
    for key, value in data.items():
        if isinstance(value, datetime):
            preview[key] = _date_to_iso(value)
        elif isinstance(value, EventType):
            preview[key] = value.value
        else:
            preview[key] = value
    return preview


def _serialize_participant_for_preview(participant: Participant) -> Dict[str, Any]:
    data = participant.model_dump(exclude_none=True)
    if "dob" in data and isinstance(participant.dob, datetime):
        data["dob"] = _date_to_iso(participant.dob)
    data["grade"] = int(participant.grade)
    if isinstance(data.get("gender"), Gender):
        data["gender"] = data["gender"].value
    return data


def _serialize_participant_event_for_preview(ep: EventParticipant) -> Dict[str, Any]:
    data = ep.model_dump(exclude_none=True)
    for key in ("travel_doc_issue_date", "travel_doc_expiry_date"):
        if key in data and isinstance(data[key], datetime):
            data[key] = _date_to_iso(data[key])
    if isinstance(data.get("transportation"), Transport):
        data["transportation"] = data["transportation"].value
    if isinstance(data.get("travel_doc_type"), DocType):
        data["travel_doc_type"] = data["travel_doc_type"].value
    if isinstance(data.get("iban_type"), IbanType):
        data["iban_type"] = data["iban_type"].value
    return data


def _merge_attendee_preview(participant: Participant, ep: EventParticipant) -> Dict[str, Any]:
    transportation = (
        ep.transportation.value
        if isinstance(ep.transportation, Transport)
        else ep.transportation
    )
    travel_doc_type = (
        ep.travel_doc_type.value
        if isinstance(ep.travel_doc_type, DocType)
        else ep.travel_doc_type
    )
    iban_type = (
        ep.iban_type.value if isinstance(ep.iban_type, IbanType) else ep.iban_type
    )

    attendee: Dict[str, Any] = {
        "pid": participant.pid,
        "name": participant.name,
        "representing_country": participant.representing_country,
        "gender": participant.gender.value if isinstance(participant.gender, Gender) else participant.gender,
        "grade": int(participant.grade),
        "dob": _date_to_iso(participant.dob),
        "pob": participant.pob,
        "birth_country": participant.birth_country,
        "citizenships": participant.citizenships or [],
        "email": participant.email,
        "phone": participant.phone,
        "diet_restrictions": participant.diet_restrictions,
        "organization": participant.organization,
        "unit": participant.unit,
        "position": participant.position,
        "rank": participant.rank,
        "intl_authority": participant.intl_authority,
        "bio_short": participant.bio_short,
    }

    attendee.update(
        {
            "event_id": ep.event_id,
            "participant_id": ep.participant_id,
            "transportation": transportation,
            "transport_other": ep.transport_other,
            "traveling_from": ep.traveling_from,
            "returning_to": ep.returning_to,
            "travel_doc_type": travel_doc_type,
            "travel_doc_issue_date": _date_to_iso(ep.travel_doc_issue_date) if isinstance(ep.travel_doc_issue_date, datetime) else "",
            "travel_doc_expiry_date": _date_to_iso(ep.travel_doc_expiry_date) if isinstance(ep.travel_doc_expiry_date, datetime) else "",
            "travel_doc_issued_by": ep.travel_doc_issued_by,
            "bank_name": ep.bank_name,
            "iban": ep.iban,
            "iban_type": iban_type,
            "swift": ep.swift,
        }
    )

    return attendee


def _load_custom_xml_objects(path: str) -> Optional[Dict[str, Any]]:
    records = _collect_custom_xml_records(path)
    if not records:
        return None

    events = []
    for rec in records.get("events", []):
        try:
            events.append(_build_event_from_record(rec))
        except Exception as exc:
            if DEBUG_PRINT:
                print(f"[CUSTOM-XML] Failed to build Event: {exc}")

    participants: List[Participant] = []
    for rec in records.get("participants", []):
        participant = _build_participant_from_record(rec)
        if participant:
            participants.append(participant)

    participant_events: List[EventParticipant] = []
    for rec in records.get("participant_events", []):
        ep = _build_participant_event_from_record(rec)
        if ep:
            participant_events.append(ep)

    if not events and not participants and not participant_events:
        return None

    return {
        "events": events,
        "event": events[0] if events else None,
        "participants": participants,
        "participant_events": participant_events,
    }

# --- ADD: lookups for ParticipantsLista and MAIN ONLINE/ParticipantsList ---
def _build_lookup_participantslista(df_positions: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    """
    Key: 'LAST|First Middle'   -> {position, phone, email}
    """
    name_col = next((c for c in df_positions.columns if "name (" in c.lower()), None)
    pos_col = next((c for c in df_positions.columns if "position" in c.lower()), None)
    phone_col = next((c for c in df_positions.columns if "phone" in c.lower()), None)
    email_col = next((c for c in df_positions.columns if "email" in c.lower()), None)
    look: Dict[str, Dict[str, str]] = {}
    if not name_col:
        return look

    for _, r in df_positions.iterrows():
        raw = _normalize(str(r.get(name_col, "")))
        key = _name_key_from_raw(raw)
        if not key:
            continue
        look[key] = {
            "position": _normalize(str(r.get(pos_col, ""))) if pos_col else "",
            "phone": _normalize(str(r.get(phone_col, ""))) if phone_col else "",
            "email": _normalize(str(r.get(email_col, ""))) if email_col else "",
        }
    return look

def _build_lookup_main_online(df_online: pd.DataFrame) -> Dict[str, Dict[str, object]]:
    """
    MAIN ONLINE → ParticipantsList table (split name columns).
    Key: 'LAST|First Middle' (and we will also try 'LAST|First' as fallback).
    """
    cols = {c.lower().strip(): c for c in df_online.columns}

    def col(label: str) -> Optional[str]:
        return cols.get(label.lower())

    look: Dict[str, Dict[str, object]] = {}
    for _, r in df_online.iterrows():
        first = _normalize(str((r.get(col("Name")) or "")))
        middle = _normalize(str((r.get(col("Middle name")) or "")))
        last = _normalize(str((r.get(col("Last name")) or "")))
        if not first and not last:
            continue
        first_middle = " ".join(part for part in [first, middle] if part).strip()
        key = _name_key(last, first_middle)
        full_name = " ".join([first, middle, last]).strip()

        gender_col = col("Gender")
        gender_raw = (str(r.get(gender_col, "")) if gender_col else "").strip()
        normalized_gender = _normalize_gender(gender_raw)
        gender = normalized_gender.value if normalized_gender is not None else gender_raw

        birth_country_raw = _normalize(str(r.get(col("Country of Birth"), "")))
        birth_country_en = translate(birth_country_raw, "en")
        birth_country_en = re.sub(r",\s*world$", "", birth_country_en, flags=re.IGNORECASE)

        travel_doc_type_col = col("Traveling document type")
        travel_doc_type_raw = (
            str(r.get(travel_doc_type_col, "")) if travel_doc_type_col else ""
        ).strip()
        travel_doc_type_translated = (
            travel_doc_type_raw, "en") if travel_doc_type_raw else ""
        travel_doc_type_value = _normalize_doc_type_label(
            travel_doc_type_translated or travel_doc_type_raw
        )

        transportation_col = col("Transportation")
        transportation_value = (
            str(r.get(transportation_col, "")) if transportation_col else ""
        ).strip()
        transport_other_col = col("Transportation (Other)")
        transport_other_value = (
            str(r.get(transport_other_col, "")) if transport_other_col else ""
        ).strip()

        iban_type_col = col("IBAN Type")
        iban_type_value = (
            str(r.get(iban_type_col, "")) if iban_type_col else ""
        ).strip()

        entry = {
            "name": full_name,
            "gender": gender,
            "dob": r.get(col("Date of Birth (DOB)")),
            "pob": _normalize(str(r.get(col("Place Of Birth (POB)"), ""))),
            "birth_country": birth_country_en,
            "citizenships": [
                _normalize(x)
                for x in re.split(r"[;,]", str(r.get(col("Citizenship(s)"), "")))
                if _normalize(x)
            ],
            "email_list": _normalize(str(r.get(col("Email address"), ""))),
            "phone_list": _normalize(str(r.get(col("Phone number"), ""))),
            "travel_doc_type": travel_doc_type_value,
            "travel_doc_number": _normalize(str(r.get(col("Traveling document number"), ""))),
            "travel_doc_issue": r.get(col("Traveling document issuance date")),
            "travel_doc_expiry": r.get(col("Traveling document expiration date")),
            "travel_doc_issued_by": translate(
                _normalize(str(r.get(col("Traveling document issued by"), ""))),
                "en",
            ),
            "transportation_declared": transportation_value,
            "transport_other": transport_other_value,
            "traveling_from_declared": _normalize(str(r.get(col("Traveling from"), ""))),
            "returning_to": _normalize(str(r.get(col("Returning to"), ""))),
            "diet_restrictions":
                _normalize(str(r.get(col("Diet restrictions"), ""))),
            "organization": translate(_normalize(str(r.get(col("Organization"), ""))), "en"),
            "unit": _normalize(str(r.get(col("Unit"), ""))),
            "rank": translate(_normalize(str(r.get(col("Rank"), ""))), "en"),
            "intl_authority": _normalize(str(r.get(col("Authority"), ""))),
            "bio_short": translate(_normalize(str(r.get(col("Short professional biography"), ""))), "en"),
            "bank_name": _normalize(str(r.get(col("Bank name"), ""))),
            "iban": _normalize(str(r.get(col("IBAN"), ""))),
            "iban_type": iban_type_value,
            "swift": _normalize(str(r.get(col("SWIFT"), ""))),
        }
        keys = [key]
        if middle and first:
            keys.append(_name_key(last, first))

        for idx, name_key in enumerate(keys):
            if not name_key:
                continue
            if idx == 0 or name_key not in look:
                look[name_key] = entry
    return look

# --- ADD: core finder for country-table columns (robust to minor header variations)
# Fixed headers used in the country tables
_EXPECTED_HEADERS = {
    "name": "Name and Last Name",
    "transport": "Travel",
    "from": "Traveling from",
    "grade": "Grade",
}

def _find_col(df: pd.DataFrame, want: str) -> Optional[str]:
    """
    Exact lookup of the expected header for the given logical key.
    Returns the header string if present, else None.
    """
    header = _EXPECTED_HEADERS.get((want or "").lower())
    if not header:
        return None

    # Fast path: exact column present
    if header in df.columns:
        return header

    # Optional tiny safety net: trim/NBSP normalize before giving up
    def norm(s: str) -> str:
        return re.sub(r"\s+", " ", (s or "").replace("\xa0", " ").strip())

    header_norm = norm(header)
    for c in df.columns:
        if norm(str(c)) == header_norm:
            return c

    # Not found
    return None


# --- ADD: the new public API for full parse (NO DB WRITES) ---
def parse_for_commit(path: str) -> dict:
    """
    Returns a dict with:
      - event: {eid, title, start_date, end_date, place, country, type, cost}
      - attendees: [ {name,
                      representing_country, transportation, traveling_from, grade,
                      position, phone, email, ...plus MAIN ONLINE fields when present} ]
    The raw attendee records (``initial_attendees``) are collected only when
    ``DEBUG_PRINT`` is enabled and otherwise omitted from the returned payload.
    """
    custom_bundle = _load_custom_xml_objects(path)
    if custom_bundle:
        event_obj: Optional[Event] = custom_bundle.get("event")
        participants: List[Participant] = custom_bundle.get("participants", [])
        participant_events: List[EventParticipant] = custom_bundle.get("participant_events", [])

        participants_by_id = {p.pid: p for p in participants}
        attendees = []
        for ep in participant_events:
            participant = participants_by_id.get(ep.participant_id)
            if not participant:
                continue
            attendees.append(_merge_attendee_preview(participant, ep))

        payload = {
            "event": event_obj.model_dump() if event_obj else {},
            "attendees": attendees,
            "objects": custom_bundle,
            "preview": {
                "event": _serialize_event_for_preview(event_obj),
                "participants": [
                    _serialize_participant_for_preview(p) for p in participants
                ],
                "participant_events": [
                    _serialize_participant_event_for_preview(ep)
                    for ep in participant_events
                ],
            },
        }

        if DEBUG_PRINT:
            print("[CUSTOM-XML] Parsed event", payload["preview"]["event"].get("eid"))
            print(
                f"[CUSTOM-XML] Participants: {len(participants)} | Participant events: {len(participant_events)}"
            )

        return payload

    # 1) Event header
    eid, title, start_date, end_date, place, country, cost = _read_event_header_block(path)

    if DEBUG_PRINT:
        print(
            "[STEP] Event header:",
            {
                "eid": eid,
                "title": title,
                "start_date": start_date,
                "end_date": end_date,
                "place": place,
                "country": country,
                "cost": cost,
            },
        )

    # 2) Tables + lookups
    tables = list_tables(path)
    idx = _index_tables(tables)

    plist = _find_table_exact(idx, "ParticipantsLista")
    ponl  = _find_table_exact(idx, "ParticipantsList")
    if not plist:
        raise RuntimeError("Required table 'ParticipantsLista' not found")
    if REQUIRE_PARTICIPANTS_LIST and not ponl:
        raise RuntimeError("Required table 'ParticipantsList' (MAIN ONLINE) not found")

    df_positions = _read_table_df(path, plist)
    df_online    = _read_table_df(path, ponl) if ponl else pd.DataFrame()

    positions_lookup = _build_lookup_participantslista(df_positions)
    online_lookup    = _build_lookup_main_online(df_online) if not df_online.empty else {}

    if DEBUG_PRINT:
        print(f"[STEP] Positions lookup entries: {len(positions_lookup)}")
        print(f"[STEP] Online lookup entries: {len(online_lookup)}")

    # 3) Collect attendees from country tables
    attendees: List[dict] = []
    initial_attendees: List[dict] = []
    for key, country_label in COUNTRY_TABLE_MAP.items():
        t = _find_table_exact(idx, key)
        if not t:
            continue
        df = _read_table_df(path, t)
        if df.empty:
            continue

        nm_col = _find_col(df, "name")
        trans_col = _find_col(df, "transport")
        from_col = _find_col(df, "from")
        grade_col = _find_col(df, "grade")

        for _, row in df.iterrows():
            raw_name = _normalize(str(row.get(nm_col, ""))) if nm_col else ""
            if not raw_name or raw_name.upper() == "TOTAL":
                continue

            transportation = _normalize(str(row.get(trans_col, ""))) if trans_col else ""
            traveling_from = _normalize(str(row.get(from_col, ""))) if from_col else ""
            grade_val = row.get(grade_col, None)
            grade = None
            if isinstance(grade_val, (int, float)) and not pd.isna(grade_val):
                try:
                    grade = int(grade_val)
                except Exception:
                    pass

            # Key for enrichment
            variants = list(_split_name_variants(raw_name))
            p_list = {}
            p_comp = {}
            for f, m, l in variants:
                key_a = _name_key(l, " ".join([f, m]).strip())
                key_b = _name_key(l, f) if f else None
                cand_list = (online_lookup.get(key_a) or (online_lookup.get(key_b) if key_b else None)) or {}
                cand_comp = (positions_lookup.get(key_a) or (positions_lookup.get(key_b) if key_b else None)) or {}
                if cand_list or cand_comp:
                    p_list, p_comp = cand_list, cand_comp
                    break

            # ✅ always fall back to {} so .get(...) is safe
            # (p_list and p_comp already default to {})

            # Determine name and display from raw country table string
            ordered = _normalize(raw_name)
            if "," in ordered:
                last_part, first_part = [x.strip() for x in ordered.split(",", 1)]
                ordered = f"{first_part} {last_part}".strip()
            base_name = ordered

            # Compose attendee record
            # Compose attendee record (country tables first)
            country_cid = _country_cid(country_label) or country_label
            transportation_value = transportation or ""
            if (not transportation_value) and p_list.get("transportation_declared"):
                transportation_value = (str(p_list.get("transportation_declared")) or "").strip()
            transportation_value = transportation_value or None

            transport_other_value = (str(p_list.get("transport_other", "")) or "").strip()
            traveling_from_value = traveling_from or p_list.get("traveling_from_declared") or ""
            grade_value = grade if grade is not None else int(Grade.NORMAL)

            base_record = {
                "name": base_name,
                "representing_country": country_cid,
                "transportation": transportation_value,
                "transport_other": transport_other_value,  # <- from country table / declared
                "traveling_from": traveling_from_value,  # <- from country table if available
                "grade": grade_value,
            }
            initial_attendees.append(base_record)

            # Start with base + ParticipantsLista (position/phone/email)
            record = {
                **base_record,
                "position": p_comp.get("position") or "",  # ParticipantsLista
                "phone": p_comp.get("phone") or "",
                "email": p_comp.get("email") or "",
            }

            # MAIN ONLINE fallback ONLY when still missing
            online = p_list or {}

            # (a) contact fields: use online only if ParticipantsLista didn’t have them
            _fill_if_missing(record, "position", online, "position_online")  # if you ever expose it
            _fill_if_missing(record, "phone", online, "phone_list")
            _fill_if_missing(record, "email", online, "email_list")

            # (b) traveling_from & transportation: country tables have priority, online as fallback
            _fill_if_missing(record, "traveling_from", online, "traveling_from_declared")
            if not record.get("transportation"):
                record["transportation"] = online.get("transportation_declared") or None

            # (c) NEVER overwrite a non-empty transport_other from base
            # (already set from country table/declared)
            # only fill if empty
            _fill_if_missing(record, "transport_other", online, "transport_other")

            # (d) fields that only MAIN ONLINE knows about (or ParticipantsLista doesn’t carry)
            birth_country_value = online.get("birth_country", "")
            birth_country_cid = resolve_birth_country_cid(
                birth_country_value,
                country_cid,
                country_label,
                lookup=_country_cid,
            )

            # Add these as new fields (they won’t exist in base/ParticipantsLista)
            record.update({
                "gender": online.get("gender", ""),
                "dob": _date_to_iso(online.get("dob")),
                "pob": online.get("pob", ""),
                "birth_country": birth_country_cid,
                "citizenships": normalize_citizenships(
                    online.get("citizenships", []),
                    lookup=_country_cid,
                ),
                "travel_doc_type": online.get("travel_doc_type"),
                "travel_doc_number": online.get("travel_doc_number", ""),
                "travel_doc_issue_date": _date_to_iso(online.get("travel_doc_issue")),
                "travel_doc_expiry_date": _date_to_iso(online.get("travel_doc_expiry")),
                "travel_doc_issued_by": online.get("travel_doc_issued_by", ""),
                "returning_to": online.get("returning_to", ""),
                "diet_restrictions": online.get("diet_restrictions", ""),
                "organization": online.get("organization", ""),
                "unit": online.get("unit", ""),
                "rank": online.get("rank", ""),
                "intl_authority": _parse_bool_value(online.get("intl_authority", "")) or False,
                "bio_short": online.get("bio_short", ""),
                "bank_name": online.get("bank_name", ""),
                "iban": online.get("iban", ""),
                "iban_type": online.get("iban_type"),
                "swift": online.get("swift", ""),
            })

            attendees.append(record)

    if DEBUG_PRINT:
        print("[STEP] Initial participant list:")
        for rec in initial_attendees:
            print("  ", rec)

    payload = {
        "event": {
            "eid": eid,
            "title": title,
            "start_date": start_date,
            "end_date": end_date,
            "place": place,
            "country": country,
            "type": "Training",
            "cost": cost,
        },
        "attendees": attendees,
    }

    payload["objects"] = None
    payload["preview"] = {
        "event": {
            "eid": eid,
            "title": title,
            "start_date": _date_to_iso(start_date),
            "end_date": _date_to_iso(end_date),
            "place": place,
            "country": country,
            "type": "Training",
            "cost": cost,
        },
        "participants": attendees,
        "participant_events": [],
    }

    if DEBUG_PRINT:
        payload["initial_attendees"] = initial_attendees

    return payload


# ============================
# String / Name helpers
# ============================

def _normalize(s: Optional[str]) -> str:
    """Normalize whitespace and coerce None to an empty string."""
    return re.sub(r"\s+", " ", (s or "").strip())


def _normalize_doc_type_label(value: object) -> str:
    """Coerce a raw travel document description to the supported enum labels."""

    if isinstance(value, DocType):
        return value.value

    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    normalized = text.lower()
    slug = re.sub(r"[^a-z0-9]+", "", normalized)
    passport_slug = re.sub(r"[^a-z0-9]+", "", DocType.passport.value.lower())
    if slug == passport_slug:
        return DocType.passport.value

    return DocType.id_card.value


def _date_to_iso(val: object) -> str:
    """
    Preview helper: format datetime to YYYY-MM-DD.
    Never returns a date object; only string or "".
    """
    if isinstance(val, datetime):
        return val.date().isoformat()
    return ""


def _norm_tablename(name: str) -> str:
    """Normalize an Excel table name to a lowercase alphanumeric key."""

    return re.sub(r"[^0-9a-zA-Z]+", "", (name or "")).lower()


def _to_app_display_name(fullname: str) -> str:
    """Convert 'First Middle Last' to 'First Middle LAST' (app display convention)."""

    name = _normalize(fullname)
    parts = name.split(" ")
    if len(parts) <= 1:
        return name
    return " ".join(parts[:-1]) + " " + parts[-1].upper()


# ============================
# Workbook / Table helpers
# ============================

def _index_tables(tables: List[TableRef]) -> Dict[str, List[TableRef]]:
    """Group tables by normalized name for quick lookup."""
    idx: Dict[str, List[TableRef]] = {}
    for t in tables:
        idx.setdefault(t.name_norm, []).append(t)
    return idx

def _find_table_exact(idx: Dict[str, List[TableRef]], desired: str) -> Optional[TableRef]:
    """Find first table with an exact normalized name match."""

    target = _norm_tablename(desired)
    group = idx.get(target)
    return group[0] if group else None

def _read_table_df(path: str, table: TableRef) -> pd.DataFrame:
    """
    Read a ListObject range (e.g., 'A4:K7') from the given sheet into a DataFrame.
    Uses the header row in the table as columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[table.sheet_title]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col, values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [_normalize(str(h)) if h is not None else "" for h in rows[0]]
    return pd.DataFrame(rows[1:], columns=header).dropna(how="all")


# ============================
# Event header parsing (A1, A2)
# ============================

def _filename_year_from_eid(filename: str) -> int:
    r"""
    From 'PFE25M2 - whatever.xlsx' infer 2025 via regex: r'PFE(\d{2})M'.
    Fallback to current UTC year if not found.
    """
    m = re.search(r"PFE(\d{2})M", filename.upper())
    return 2000 + int(m.group(1)) if m else datetime.now(UTC).year

def _parse_event_header(a1: str, a2: str, year: int):
    """
    A1: 'PFE25M2 TITLE OF THE EVENT' -> (eid, title)
    A2: 'JUNE 23 - 27 - Opatija, CROATIA' -> (start_date, end_date, place, country)
    Month for end date assumed same as start (per your spec).
    """
    a1 = _normalize(a1)
    sp = a1.find(" ")
    eid, title = (a1, "") if sp == -1 else (a1[:sp], a1[sp + 1:])

    a2 = _normalize(a2)
    parts = [p.strip() for p in a2.split(" - ")]
    start_date = end_date = None
    location = ""
    if len(parts) >= 3:
        month_and_start, end_day_str, location = parts[0], parts[1], parts[2]
        m = re.match(r"([A-Z]+)\s+(\d{1,2})", month_and_start.upper())
        if m:
            month_num = MONTHS.get(m.group(1))
            start_day = int(m.group(2))
            if month_num:
                end_day = int(re.sub(r"\D", "", end_day_str))
                start_date = datetime(year, month_num, start_day, tzinfo=UTC)
                end_date = datetime(year, month_num, end_day, tzinfo=UTC)

    place = location
    country_value: str | None = None
    if location:
        loc_parts = [p.strip() for p in location.split(",", 1)]
        place = loc_parts[0]
        raw_country = loc_parts[1] if len(loc_parts) > 1 else ""
        if raw_country:
            normalized_country = translate(_normalize(raw_country), "en")
            lookup = _country_cid(normalized_country) or _country_cid(normalized_country.title())
            country_value = lookup or normalized_country

    return eid, title, start_date, end_date, place, country_value


# ============================
# DB helpers (read-only)
# ============================

def _country_cid(name: str) -> Optional[str]:
    """Return the country ``cid`` for ``name``, or ``None`` if not found."""
    if not name:
        return None
    try:
        doc = mongodb.collection('countries').find_one({'country': name})
        return doc.get('cid') if doc else None
    except Exception:
        # Avoid raising during preview; treat missing/DB errors as not found
        return None

def _participant_exists(name_display: str, country_name: str):
    """Existence check by normalized app name and representing country CID."""
    q = {"name": _to_app_display_name(name_display)}
    cid = _country_cid(country_name)
    if cid is not None:
        q["representing_country"] = cid
    try:
        doc = mongodb.collection('participants').find_one(q)
        return (doc is not None), (doc or {})
    except Exception:
        return False, {}


# ============================
# Public API
# ============================

def validate_excel_file_for_import(path: str) -> tuple[bool, list[str], dict]:
    """
    Validate that:
      - Sheet 'Participants' exists
      - A1 (eid+title) and A2 (dates+location) present (non-empty)
      - B15 (cost) present (non-empty)
      - Table 'ParticipantsLista' exists (any sheet)
      - At least one of the country tables exists (any sheet)
    Returns: (ok, missing_list, tables_info_dict)
    """
    custom_bundle = _load_custom_xml_objects(path)
    if custom_bundle:
        event_obj: Optional[Event] = custom_bundle.get("event")
        participants = custom_bundle.get("participants", [])
        participant_events = custom_bundle.get("participant_events", [])
        seen = {
            "custom_xml": True,
            "events": len(custom_bundle.get("events", [])),
            "participants": len(participants),
            "participant_events": len(participant_events),
        }
        if event_obj:
            seen["event_eid"] = event_obj.eid
        return True, [], seen

    missing: list[str] = []

    # 1) A1/A2 on "Participants"
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Participants" not in wb.sheetnames:
        missing.append("Sheet 'Participants'")
        return False, missing, {}
    if "COST Overview" not in wb.sheetnames:
        missing.append("Sheet 'COST Overview'")
        return False, missing, {}
    ws = wb["Participants"]
    wws = wb["COST Overview"]
    a1 = (ws["A1"].value or "").strip()
    a2 = (ws["A2"].value or "").strip()
    cost_overview_b15 = str(wws["B15"].value or "").strip()
    if not a1:
        missing.append("Participants!A1 (eid + title)")
    if not a2:
        missing.append("Participants!A2 (dates + location)")
    if not cost_overview_b15:
        missing.append("Cost Overview!B15 (Total Cost)")

    # 2) Tables via inspector (cross-sheet, ZIP-safe)
    tables = list_tables(path)
    idx = _index_tables(tables)

    # ParticipantsLista present?
    if not _find_table_exact(idx, "ParticipantsLista"):
        missing.append("Table 'ParticipantsLista'")

    # At least one country table present?
    if not any(_find_table_exact(idx, k) for k in COUNTRY_TABLE_MAP.keys()):
        missing.append("At least one country table (tableAlb, tableBih, tableCro, tableKos, tableMne, tableNmk, tableSer)")

    ok = len(missing) == 0

    # For visibility return a compact dict of what we saw
    seen = {t.name_norm: (t.name, t.sheet_title, t.ref) for t in tables}

    if DEBUG_PRINT:
        print("[VALIDATE] OK:", ok)
        if missing:
            print("[VALIDATE] Missing:", missing)
        print("[VALIDATE] Tables seen:", seen)

    return ok, missing, seen


def inspect_and_preview_uploaded(path: str) -> None:
    """
    NO WRITES.
    - Logs EVENT NEW/EXIST (from A1/A2 on 'Participants' sheet)
    - Logs ATTENDEES from any present country tables
      (marks new with '*', includes grade, country, and position from ParticipantsLista)
    """
    # Event header
    eid, title, start_date, end_date, place, country, cost = _read_event_header_block(path)

    # Event exist check (read-only)
    existing = mongodb.collection('events').find_one({"eid": eid})
    if existing:
        existing_start = existing.get('start_date') or existing.get('dateFrom')
        existing_place = existing.get('place') or existing.get('location', '')
        existing_country = existing.get('country')
        print(
            f"[EVENT] EXIST {eid}  title='{existing.get('title','')}' "
            f"start_date={existing_start} place='{existing_place}' country='{existing_country}'"
        )
    else:
        print(
            f"[EVENT] NEW   {eid}  title='{title}' start_date={start_date} "
            f"end_date={end_date} place='{place}' country='{country}'"
        )

    # Tables + positions
    tables = list_tables(path)
    idx = _index_tables(tables)

    plist = _find_table_exact(idx, "ParticipantsLista")
    if not plist:
        raise RuntimeError("Required table 'ParticipantsLista' not found (any sheet)")

    df_positions = _read_table_df(path, plist)
    positions_lookup_full = _build_lookup_participantslista(df_positions)

    print("[ATTENDEES]")

    # Iterate any present country tables
    for key, country_label in COUNTRY_TABLE_MAP.items():
        t = _find_table_exact(idx, key)
        if not t:
            continue

        df = _read_table_df(path, t)
        if df.empty:
            continue

        nm_col = next((c for c in df.columns if "name" in c.lower()), None)
        grade_col = next((c for c in df.columns if "grade" in c.lower()), None)

        for _, row in df.iterrows():
            raw_name = _normalize(str(row.get(nm_col, ""))) if nm_col else ""
            if not raw_name:
                continue
            grade = _normalize(str(row.get(grade_col, ""))) if grade_col else ""

            # Build key to lookup position: "LAST|First Middle"
            key_lookup = _name_key_from_raw(raw_name)
            pos = positions_lookup_full.get(key_lookup, {}).get("position", "")

            exists, doc = _participant_exists(raw_name, country_label)
            norm = _to_app_display_name(raw_name)
            star = "*" if not exists else " "
            pid = doc.get("pid", "NEW")

            print(
                f"{star} {'NEW' if star=='*' else 'EXIST'} {pid:>6}  {norm}  ({grade}, {country_label})  "
                f"{'pos='+pos if pos else ''}"
            )


def _name_key_from_raw(raw_display: str) -> str:
    s = _normalize(raw_display)
    if not s:
        return ""
    if "," in s:
        last, first = [x.strip() for x in s.split(",", 1)]
    else:
        parts = s.split()
        last = parts[-1] if len(parts) > 1 else s
        first = " ".join(parts[:-1]) if len(parts) > 1 else ""
    return _name_key(last, first)

def _read_event_header_block(path: str) -> tuple[str, str, datetime, datetime, str, Optional[str], Optional[float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Participants" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Participants' not found")
    if "COST Overview" not in wb.sheetnames:
        raise RuntimeError("Sheet 'COST Overview' not found")

    ws = wb["Participants"]
    a1 = ws["A1"].value or ""
    a2 = ws["A2"].value or ""
    year = _filename_year_from_eid(os.path.basename(path))
    eid, title, start_date, end_date, place, country = _parse_event_header(a1, a2, year)

    wws = wb["COST Overview"]
    cost_overview_b15 = str(wws["B15"].value or "").strip()
    cost = float(cost_overview_b15) if cost_overview_b15 else None
    return eid, title, start_date, end_date, place, country, cost

def _fill_if_missing(dst: dict, key: str, src: dict, src_key: Optional[str] = None) -> None:
    """If dst[key] is falsy (None/''/0/[]), copy from src[src_key or key] if present and truthy."""
    k = src_key or key
    if not dst.get(key) and src.get(k):
        dst[key] = src[k]
